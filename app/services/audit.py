# services/audit.py - Serviço de Auditoria

from typing import Optional, Dict, Any
from sqlalchemy.orm import Session
from datetime import datetime, timezone

from app.models_audit import (
    CustomerAuditLog,
    ProductAuditLog,
    SystemUserAuditLog,
    SaleAuditLog,
    AuditAction
)


class AuditService:
    """
    Serviço centralizado para registrar ações de auditoria.
    Facilita o uso e garante consistência.
    """

    def __init__(self, db: Session):
        self.db = db

    # ============================================
    # MÉTODOS PARA CLIENTES
    # ============================================

    def log_customer_action(
        self,
        customer_id: int,
        action: AuditAction,
        created_by_id: int,
        old_values: Optional[Dict[str, Any]] = None,
        new_values: Optional[Dict[str, Any]] = None,
        description: Optional[str] = None
    ):
        """
        Registra uma ação relacionada a cliente.

        Exemplo de uso:
            audit.log_customer_action(
                customer_id=123,
                action=AuditAction.UPDATE,
                created_by_id=current_user.id,
                old_values={"nome": "João", "saldo": 10.0},
                new_values={"nome": "João Silva", "saldo": 10.0},
                description="Atualizou nome do cliente"
            )
        """
        log = CustomerAuditLog(
            customer_id=customer_id,
            action=action,
            created_by_id=created_by_id,
            old_values=old_values,
            new_values=new_values,
            description=description
        )
        self.db.add(log)
        self.db.commit()
        return log

    def get_customer_history(
        self,
        customer_id: int,
        limit: int = 50
    ):
        """
        Busca histórico de um cliente específico.
        ⚡ Rápido graças ao índice idx_customer_date
        """
        from app.models import SystemUser

        logs = self.db.query(CustomerAuditLog)\
            .filter(CustomerAuditLog.customer_id == customer_id)\
            .order_by(CustomerAuditLog.created_at.desc())\
            .limit(limit)\
            .all()

        # Popular created_by_username
        for log in logs:
            if log.created_by_id:
                user = self.db.query(SystemUser).filter(SystemUser.id == log.created_by_id).first()
                log.created_by_username = user.username if user else None

        return logs

    # ============================================
    # MÉTODOS PARA PRODUTOS
    # ============================================

    def log_product_action(
        self,
        produto_id: int,
        action: AuditAction,
        created_by_id: int,
        old_values: Optional[Dict[str, Any]] = None,
        new_values: Optional[Dict[str, Any]] = None,
        description: Optional[str] = None
    ):
        """
        Registra uma ação relacionada a produto.

        Exemplo:
            audit.log_product_action(
                produto_id=456,
                action=AuditAction.PRICE_CHANGE,
                created_by_id=current_user.id,
                old_values={"valor": 5.0},
                new_values={"valor": 7.0},
                description="Ajuste de preço"
            )
        """
        log = ProductAuditLog(
            produto_id=produto_id,
            action=action,
            created_by_id=created_by_id,
            old_values=old_values,
            new_values=new_values,
            description=description
        )
        self.db.add(log)
        self.db.commit()
        return log

    def get_product_history(
        self,
        produto_id: int,
        limit: int = 50
    ):
        """Busca histórico de um produto específico"""
        from app.models import SystemUser

        logs = self.db.query(ProductAuditLog)\
            .filter(ProductAuditLog.produto_id == produto_id)\
            .order_by(ProductAuditLog.created_at.desc())\
            .limit(limit)\
            .all()

        # Popular created_by_username
        for log in logs:
            if log.created_by_id:
                user = self.db.query(SystemUser).filter(SystemUser.id == log.created_by_id).first()
                log.created_by_username = user.username if user else None

        return logs

    def get_price_changes(self, produto_id: int):
        """Busca apenas mudanças de preço de um produto"""
        from app.models import SystemUser

        logs = self.db.query(ProductAuditLog)\
            .filter(
                ProductAuditLog.produto_id == produto_id,
                ProductAuditLog.action == AuditAction.PRICE_CHANGE
            )\
            .order_by(ProductAuditLog.created_at.desc())\
            .all()

        # Popular created_by_username
        for log in logs:
            if log.created_by_id:
                user = self.db.query(SystemUser).filter(SystemUser.id == log.created_by_id).first()
                log.created_by_username = user.username if user else None

        return logs

    # ============================================
    # MÉTODOS PARA USUÁRIOS
    # ============================================

    def log_user_action(
        self,
        user_id: int,
        action: AuditAction,
        created_by_id: int,
        old_values: Optional[Dict[str, Any]] = None,
        new_values: Optional[Dict[str, Any]] = None,
        description: Optional[str] = None
    ):
        """
        Registra uma ação relacionada a usuário do sistema.

        Exemplo:
            audit.log_user_action(
                user_id=789,
                action=AuditAction.DEACTIVATE,
                created_by_id=current_admin.id,
                old_values={"is_active": True},
                new_values={"is_active": False},
                description="Usuário desativado por inatividade"
            )
        """
        log = SystemUserAuditLog(
            user_id=user_id,
            action=action,
            created_by_id=created_by_id,
            old_values=old_values,
            new_values=new_values,
            description=description
        )
        self.db.add(log)
        self.db.commit()
        return log

    def get_user_history(
        self,
        user_id: int,
        limit: int = 50
    ):
        """Busca histórico de um usuário específico"""
        from app.models import SystemUser

        logs = self.db.query(SystemUserAuditLog)\
            .filter(SystemUserAuditLog.user_id == user_id)\
            .order_by(SystemUserAuditLog.created_at.desc())\
            .limit(limit)\
            .all()

        # Popular created_by_username
        for log in logs:
            if log.created_by_id:
                user = self.db.query(SystemUser).filter(SystemUser.id == log.created_by_id).first()
                log.created_by_username = user.username if user else None

        return logs

    # ============================================
    # MÉTODOS PARA VENDAS
    # ============================================

    def log_sale_create(
        self,
        sale_id: int,
        customer_id: int,
        customer_nome: str,
        customer_nickname: str,
        total_amount: float,
        items_count: int,
        items: list,
        old_customer_balance: float,
        new_customer_balance: float,
        created_by_id: int
    ):
        """
        Registra a criação de uma venda.
        """
        log = SaleAuditLog(
            sale_id=sale_id,
            action=AuditAction.CREATE,
            created_by_id=created_by_id,
            old_values={},
            new_values={
                "customer_id": customer_id,
                "customer_nome": customer_nome,
                "customer_nickname": customer_nickname,
                "total_amount": total_amount,
                "items_count": items_count,
                "items": items,
                "old_customer_balance": old_customer_balance,
                "new_customer_balance": new_customer_balance
            },
            description=f"Sale create - Sale ID: {sale_id}"
        )
        self.db.add(log)
        self.db.commit()
        return log

    def log_sale_cancel(
        self,
        sale_id: int,
        customer_id: int,
        customer_nome: str,
        total_amount: float,
        reason: str,
        old_customer_balance: float,
        new_customer_balance: float,
        created_by_id: int
    ):
        """
        Registra o cancelamento de uma venda.
        """
        log = SaleAuditLog(
            sale_id=sale_id,
            action=AuditAction.DELETE,  # Cancelar = DELETE lógico
            created_by_id=created_by_id,
            old_values={
                "total_amount": total_amount,
                "customer_balance": old_customer_balance
            },
            new_values={
                "is_cancelled": True,
                "cancellation_reason": reason,
                "customer_balance": new_customer_balance,
                "refunded_amount": total_amount
            },
            description=f"Sale cancelled - Sale ID: {sale_id}, Reason: {reason}"
        )
        self.db.add(log)
        self.db.commit()
        return log

    def get_sale_history(
        self,
        sale_id: int,
        limit: int = 50
    ):
        """Busca histórico de uma venda específica"""
        from app.models import SystemUser

        logs = self.db.query(SaleAuditLog)\
            .filter(SaleAuditLog.sale_id == sale_id)\
            .order_by(SaleAuditLog.created_at.desc())\
            .limit(limit)\
            .all()

        # Popular created_by_username
        for log in logs:
            if log.created_by_id:
                user = self.db.query(SystemUser).filter(SystemUser.id == log.created_by_id).first()
                log.created_by_username = user.username if user else None

        return logs

    # ============================================
    # MÉTODO GENÉRICO (para entidades não específicas como Sale)
    # ============================================

    def log_action(
        self,
        user_id: int,
        action: str,
        entity_type: str,
        entity_id: int,
        details: Optional[Dict[str, Any]] = None
    ):
        """
        Método genérico para registrar ações de qualquer tipo de entidade.
        Usa a tabela apropriada baseada no entity_type.

        Exemplo:
            audit.log_action(
                user_id=current_user.id,
                action="CREATE",
                entity_type="Sale",
                entity_id=123,
                details={
                    "customer_id": 45,
                    "total_amount": 15.50,
                    "items_count": 3
                }
            )
        """
        # Converter action string para AuditAction enum
        try:
            if hasattr(AuditAction, action.upper()):
                action_enum = getattr(AuditAction, action.upper())
            else:
                action_enum = AuditAction.UPDATE  # fallback
        except (KeyError, AttributeError):
            action_enum = AuditAction.UPDATE

        # Normalizar entity_type (case-insensitive)
        entity_type_normalized = entity_type.lower().strip()

        # Criar descrição baseada no entity_type e action
        description = f"{entity_type} {action.lower()}"
        if details:
            description += f" - {entity_type} ID: {entity_id}"

        # Usar tabela apropriada baseada no entity_type
        # IMPORTANTE: Verificar o tipo de entidade de forma robusta
        if entity_type_normalized == "sale":
            # ✅ VENDAS -> SaleAuditLog
            log = SaleAuditLog(
                sale_id=entity_id,
                action=action_enum,
                created_by_id=user_id,
                old_values={},
                new_values=details,
                description=description
            )
        elif entity_type_normalized == "customer":
            # ✅ CLIENTES -> CustomerAuditLog
            log = CustomerAuditLog(
                customer_id=entity_id,
                action=action_enum,
                created_by_id=user_id,
                old_values={},
                new_values=details,
                description=description
            )
        elif entity_type_normalized == "product" or entity_type_normalized == "produto":
            # ✅ PRODUTOS -> ProductAuditLog
            log = ProductAuditLog(
                produto_id=entity_id,
                action=action_enum,
                created_by_id=user_id,
                old_values={},
                new_values=details,
                description=description
            )
        elif entity_type_normalized == "user" or entity_type_normalized == "systemuser":
            # ✅ USUÁRIOS -> SystemUserAuditLog
            log = SystemUserAuditLog(
                user_id=entity_id,
                action=action_enum,
                created_by_id=user_id,
                old_values={},
                new_values=details,
                description=description
            )
        else:
            # ⚠️ TIPO DESCONHECIDO - Log de erro e usar fallback
            # Mas NÃO salvar em SystemUserAuditLog com entity_id errado
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(
                f"Tipo de entidade desconhecido: '{entity_type}' (normalizado: '{entity_type_normalized}'). "
                f"Registrando em SystemUserAuditLog com old_values contendo metadados."
            )

            # Usar SystemUserAuditLog mas com user_id = created_by_id
            # e colocar os dados reais em old_values/new_values
            log = SystemUserAuditLog(
                user_id=user_id,  # ← IMPORTANTE: usar user_id, não entity_id
                action=action_enum,
                created_by_id=user_id,
                old_values={
                    "entity_type": entity_type,
                    "entity_id": entity_id,
                    "warning": "Tipo de entidade não reconhecido"
                },
                new_values=details,
                description=f"[UNKNOWN TYPE: {entity_type}] {description}"
            )

        self.db.add(log)
        self.db.commit()
        return log

    # ============================================
    # RELATÓRIOS GLOBAIS
    # ============================================

    def get_user_activity(
        self,
        user_id: int,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ):
        """
        Busca todas as ações realizadas por um usuário.
        ⚡ Usa índices idx_user_action de cada tabela
        """
        # Buscar logs de clientes criados/modificados por este usuário
        customer_query = self.db.query(CustomerAuditLog)\
            .filter(CustomerAuditLog.created_by_id == user_id)

        if start_date:
            customer_query = customer_query.filter(CustomerAuditLog.created_at >= start_date)
        if end_date:
            customer_query = customer_query.filter(CustomerAuditLog.created_at <= end_date)

        customer_logs = customer_query.all()

        # Buscar logs de produtos criados/modificados por este usuário
        product_query = self.db.query(ProductAuditLog)\
            .filter(ProductAuditLog.created_by_id == user_id)

        if start_date:
            product_query = product_query.filter(ProductAuditLog.created_at >= start_date)
        if end_date:
            product_query = product_query.filter(ProductAuditLog.created_at <= end_date)

        product_logs = product_query.all()

        # Buscar logs de usuários criados/modificados por este usuário
        user_query = self.db.query(SystemUserAuditLog)\
            .filter(SystemUserAuditLog.created_by_id == user_id)

        if start_date:
            user_query = user_query.filter(SystemUserAuditLog.created_at >= start_date)
        if end_date:
            user_query = user_query.filter(SystemUserAuditLog.created_at <= end_date)

        user_logs = user_query.all()

        # Buscar logs de vendas criadas/canceladas por este usuário
        sale_query = self.db.query(SaleAuditLog)\
            .filter(SaleAuditLog.created_by_id == user_id)

        if start_date:
            sale_query = sale_query.filter(SaleAuditLog.created_at >= start_date)
        if end_date:
            sale_query = sale_query.filter(SaleAuditLog.created_at <= end_date)

        sale_logs = sale_query.all()

        # Combinar e ordenar
        all_logs = customer_logs + product_logs + user_logs + sale_logs
        all_logs.sort(key=lambda x: x.created_at, reverse=True)

        return all_logs

    def get_recent_activity(self, limit: int = 100):
        """
        Busca atividade recente do sistema (todas as entidades).
        Útil para dashboard de administração.
        """
        # Buscar os mais recentes de cada tipo
        customer_logs = self.db.query(CustomerAuditLog)\
            .order_by(CustomerAuditLog.created_at.desc())\
            .limit(limit)\
            .all()

        product_logs = self.db.query(ProductAuditLog)\
            .order_by(ProductAuditLog.created_at.desc())\
            .limit(limit)\
            .all()

        user_logs = self.db.query(SystemUserAuditLog)\
            .order_by(SystemUserAuditLog.created_at.desc())\
            .limit(limit)\
            .all()

        sale_logs = self.db.query(SaleAuditLog)\
            .order_by(SaleAuditLog.created_at.desc())\
            .limit(limit)\
            .all()

        # Combinar e ordenar
        all_logs = customer_logs + product_logs + user_logs + sale_logs
        all_logs.sort(key=lambda x: x.created_at, reverse=True)

        return all_logs[:limit]

    def search_logs_filtered(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        user_id: Optional[int] = None,
        action: Optional[str] = None,
        entity_type: Optional[str] = None,
        entity_id: Optional[int] = None,
        entity_name: Optional[str] = None,
        search: Optional[str] = None,
        limit: int = 100,
        offset: int = 0
    ):
        """
        Busca logs de auditoria com filtros avançados.

        Filtros disponíveis:
        - start_date/end_date: Período
        - user_id: Usuário que realizou a ação
        - action: Tipo de ação (CREATE, UPDATE, DELETE, etc)
        - entity_type: Módulo (customer, product, user, sale)
        - entity_id: ID específico da entidade
        - entity_name: Nome da entidade (busca parcial por nome do cliente, produto, usuário)
        - search: Busca por texto na descrição
        - limit/offset: Paginação

        Retorna: (logs, total_count)
        """
        from app.models import SystemUser, Customers, Produto
        from sqlalchemy import or_

        # Configuração das entidades - Define tudo em um único lugar
        entity_configs = {
            'customer': {
                'model': CustomerAuditLog,
                'entity_id_field': 'customer_id',
                'entity_table': Customers,
                'entity_name_fields': ['nome', 'nickname'],
                'display_name': 'Cliente'
            },
            'product': {
                'model': ProductAuditLog,
                'entity_id_field': 'produto_id',
                'entity_table': Produto,
                'entity_name_fields': ['nome'],
                'display_name': 'Produto'
            },
            'user': {
                'model': SystemUserAuditLog,
                'entity_id_field': 'user_id',
                'entity_table': SystemUser,
                'entity_name_fields': ['username'],
                'display_name': 'Usuário'
            },
            'sale': {
                'model': SaleAuditLog,
                'entity_id_field': 'sale_id',
                'entity_table': None,
                'entity_name_fields': None,
                'display_name': 'Venda'
            }
        }

        # Determinar quais entidades consultar
        entities_to_query = [entity_type.lower()] if entity_type else entity_configs.keys()

        # Função auxiliar para aplicar filtros comuns (elimina duplicação)
        def apply_common_filters(query, model):
            """Aplica filtros que são comuns a todas as entidades"""
            if start_date:
                query = query.filter(model.created_at >= start_date)
            if end_date:
                query = query.filter(model.created_at <= end_date)
            if user_id:
                query = query.filter(model.created_by_id == user_id)
            if action:
                try:
                    action_enum = AuditAction[action.upper()]
                    query = query.filter(model.action == action_enum)
                except KeyError:
                    pass  # Ignora ações inválidas
            if search:
                query = query.filter(
                    or_(
                        model.description.ilike(f"%{search}%"),
                        model.new_values.astext.ilike(f"%{search}%")
                    )
                )
            return query

        # Coletar todos os logs
        all_logs = []
        total_count = 0

        for entity_type_name in entities_to_query:
            if entity_type_name not in entity_configs:
                continue

            config = entity_configs[entity_type_name]
            model = config['model']

            # Iniciar query
            query = self.db.query(model)

            # Aplicar filtros comuns
            query = apply_common_filters(query, model)

            # Aplicar filtro de entity_id
            if entity_id:
                query = query.filter(getattr(model, config['entity_id_field']) == entity_id)

            # Aplicar filtro de entity_name (com JOIN se necessário)
            if entity_name and config['entity_table']:
                entity_table = config['entity_table']
                entity_id_col = getattr(model, config['entity_id_field'])

                query = query.join(entity_table, entity_id_col == entity_table.id)

                # Criar condições OR para múltiplos campos de nome
                name_conditions = [
                    getattr(entity_table, field).ilike(f"%{entity_name}%")
                    for field in config['entity_name_fields']
                ]
                query = query.filter(or_(*name_conditions))

            # Contar total
            count = query.count()
            total_count += count

            # Buscar logs
            logs = query.order_by(model.created_at.desc()).all()

            # Enriquecer dados
            for log in logs:
                self._enrich_log(log, entity_type_name, config)
                all_logs.append(log)

        # Ordenar por data (mais recente primeiro)
        all_logs.sort(key=lambda x: x.created_at, reverse=True)

        # Aplicar paginação
        paginated_logs = all_logs[offset:offset + limit]

        return paginated_logs, total_count

    def _enrich_log(self, log, entity_type_name: str, config: dict):
        """
        Enriquece um log com informações adicionais (método auxiliar interno).
        Separa a lógica de enriquecimento para manter o código limpo.
        """
        from app.models import SystemUser

        # Adicionar tipo e ID da entidade
        log.entity_type = entity_type_name
        log.entity_id = getattr(log, config['entity_id_field'])

        # Buscar nome da entidade
        if config['entity_table']:
            entity = self.db.query(config['entity_table'])\
                .filter(config['entity_table'].id == log.entity_id)\
                .first()

            if entity:
                # Pegar o primeiro campo de nome disponível
                name_field = config['entity_name_fields'][0]
                log.entity_name = getattr(entity, name_field)
            else:
                log.entity_name = f"{config['display_name']} #{log.entity_id}"
        else:
            log.entity_name = f"{config['display_name']} #{log.entity_id}"

        # Buscar nome do usuário que realizou a ação
        created_by = self.db.query(SystemUser)\
            .filter(SystemUser.id == log.created_by_id)\
            .first()
        log.created_by_username = created_by.username if created_by else "Sistema"

    def get_sale_history(self, sale_id: int):
        """
        Busca histórico de auditoria de uma venda específica.
        """
        logs = self.db.query(SaleAuditLog)\
            .filter(SaleAuditLog.sale_id == sale_id)\
            .order_by(SaleAuditLog.created_at.desc())\
            .all()

        return logs

    # ============================================
    # EXPORTAÇÃO DE LOGS
    # ============================================

    def _get_filename(
        self,
        base_name: str,
        extension: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        user_id: Optional[int] = None,
        action: Optional[str] = None,
        entity_type: Optional[str] = None,
        event_name: Optional[str] = None
    ) -> str:
        """
        Gera nome de arquivo com metadados (usado por todos os formatos).
        Inclui o nome do evento se disponível.
        """
        from datetime import datetime as dt

        filename_parts = []

        # Adicionar nome do evento se disponível
        if event_name:
            filename_parts.append(event_name)

        filename_parts.append(base_name)

        # Adicionar timestamp
        timestamp = dt.now().strftime('%Y-%m-%d_%H-%M-%S')
        filename_parts.append(timestamp)

        # Adicionar filtros ao nome (para rastreabilidade)
        if user_id:
            filename_parts.append(f'user_{user_id}')
        if entity_type:
            filename_parts.append(entity_type)
        if action:
            filename_parts.append(action.lower())
        if start_date:
            filename_parts.append(f'from_{start_date.strftime("%Y%m%d")}')
        if end_date:
            filename_parts.append(f'to_{end_date.strftime("%Y%m%d")}')

        filename = '_'.join(filename_parts) + extension
        return filename

    def export_logs_to_csv(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        user_id: Optional[int] = None,
        action: Optional[str] = None,
        entity_type: Optional[str] = None,
        entity_id: Optional[int] = None,
        entity_name: Optional[str] = None,
        search: Optional[str] = None,
        limit: int = 10000,
        event_name: Optional[str] = None
    ):
        """
        Exporta logs de auditoria para CSV.

        **Vantagens do CSV:**
        - Universal (Excel, Google Sheets, LibreOffice)
        - Leve e rápido
        - Fácil de processar programaticamente

        Retorna: (csv_content: str, filename: str, total: int)
        """
        import csv
        import io

        # Buscar logs com os filtros
        logs, total = self.search_logs_filtered(
            start_date=start_date,
            end_date=end_date,
            user_id=user_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            search=search,
            limit=limit,
            offset=0
        )

        # Criar CSV em memória
        output = io.StringIO()
        writer = csv.writer(output)

        # Adicionar informações de filtros no topo
        writer.writerow(['RELATÓRIO DE LOGS DE AUDITORIA'])
        if event_name:
            writer.writerow(['Evento:', event_name])
        writer.writerow(['Gerado em:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow(['Total de registros:', total])
        writer.writerow([])  # Linha em branco

        # Filtros aplicados
        writer.writerow(['FILTROS APLICADOS:'])
        if start_date:
            writer.writerow(['Data inicial:', start_date.strftime('%Y-%m-%d %H:%M:%S')])
        if end_date:
            writer.writerow(['Data final:', end_date.strftime('%Y-%m-%d %H:%M:%S')])
        if user_id:
            writer.writerow(['ID do usuário:', user_id])
        if action:
            writer.writerow(['Tipo de ação:', action])
        if entity_type:
            writer.writerow(['Módulo:', entity_type.upper()])
        if entity_id:
            writer.writerow(['ID da entidade:', entity_id])
        if entity_name:
            writer.writerow(['Nome da entidade:', entity_name])
        if search:
            writer.writerow(['Busca por texto:', search])

        if not any([start_date, end_date, user_id, action, entity_type, entity_id, entity_name, search]):
            writer.writerow(['Nenhum filtro aplicado (todos os registros)'])

        writer.writerow([])  # Linha em branco
        writer.writerow([])  # Linha em branco

        # Cabeçalho da tabela de dados
        writer.writerow([
            'ID',
            'Data/Hora',
            'Módulo',
            'Entidade',
            'ID da Entidade',
            'Ação',
            'Usuário',
            'ID do Usuário',
            'Valores Antigos',
            'Valores Novos',
            'Descrição'
        ])

        # Dados
        for log in logs:
            writer.writerow([
                log.id,
                log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                log.entity_type.upper() if hasattr(log, 'entity_type') else '',
                getattr(log, 'entity_name', ''),
                getattr(log, 'entity_id', ''),
                log.action.value if hasattr(log.action, 'value') else str(log.action),
                getattr(log, 'created_by_username', ''),
                log.created_by_id,
                str(log.old_values) if log.old_values else '',
                str(log.new_values) if log.new_values else '',
                log.description or ''
            ])

        # Gerar nome do arquivo
        filename = self._get_filename(
            'audit_logs', '.csv',
            start_date, end_date, user_id, action, entity_type, event_name
        )

        csv_content = output.getvalue()
        output.close()

        return csv_content, filename, total

    def export_logs_to_json(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        user_id: Optional[int] = None,
        action: Optional[str] = None,
        entity_type: Optional[str] = None,
        entity_id: Optional[int] = None,
        entity_name: Optional[str] = None,
        search: Optional[str] = None,
        limit: int = 10000,
        event_name: Optional[str] = None
    ):
        """
        Exporta logs de auditoria para JSON.

        **Vantagens do JSON:**
        - Ideal para APIs e integração com sistemas
        - Preserva estrutura de dados complexa
        - Fácil de parsear programaticamente
        - Suporta aninhamento de objetos

        Retorna: (json_content: str, filename: str, total: int)
        """
        import json

        # Buscar logs com os filtros
        logs, total = self.search_logs_filtered(
            start_date=start_date,
            end_date=end_date,
            user_id=user_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            search=search,
            limit=limit,
            offset=0
        )

        # Preparar dados para JSON
        logs_data = []
        for log in logs:
            logs_data.append({
                'id': log.id,
                'created_at': log.created_at.isoformat(),
                'entity': {
                    'type': getattr(log, 'entity_type', ''),
                    'id': getattr(log, 'entity_id', None),
                    'name': getattr(log, 'entity_name', '')
                },
                'action': log.action.value if hasattr(log.action, 'value') else str(log.action),
                'user': {
                    'id': log.created_by_id,
                    'username': getattr(log, 'created_by_username', '')
                },
                'changes': {
                    'old': log.old_values,
                    'new': log.new_values
                },
                'description': log.description or ''
            })

        # Criar estrutura JSON completa
        # Gerar descrição legível dos filtros
        filters_description = []
        if start_date:
            filters_description.append(f"Data inicial: {start_date.strftime('%Y-%m-%d %H:%M:%S')}")
        if end_date:
            filters_description.append(f"Data final: {end_date.strftime('%Y-%m-%d %H:%M:%S')}")
        if user_id:
            filters_description.append(f"ID do usuário: {user_id}")
        if action:
            filters_description.append(f"Tipo de ação: {action}")
        if entity_type:
            filters_description.append(f"Módulo: {entity_type.upper()}")
        if entity_id:
            filters_description.append(f"ID da entidade: {entity_id}")
        if entity_name:
            filters_description.append(f"Nome da entidade: {entity_name}")
        if search:
            filters_description.append(f"Busca por texto: {search}")

        if not filters_description:
            filters_description.append("Nenhum filtro aplicado (todos os registros)")

        export_data = {
            'metadata': {
                'event_name': event_name,
                'exported_at': datetime.now().isoformat(),
                'total_records': total,
                'filters_applied': filters_description,
                'filters': {
                    'start_date': start_date.isoformat() if start_date else None,
                    'end_date': end_date.isoformat() if end_date else None,
                    'user_id': user_id,
                    'action': action,
                    'entity_type': entity_type,
                    'entity_id': entity_id,
                    'entity_name': entity_name,
                    'search': search
                }
            },
            'logs': logs_data
        }

        # Gerar JSON
        json_content = json.dumps(export_data, indent=2, ensure_ascii=False)

        # Gerar nome do arquivo
        filename = self._get_filename(
            'audit_logs', '.json',
            start_date, end_date, user_id, action, entity_type, event_name
        )

        return json_content, filename, total

    def export_logs_to_excel(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        user_id: Optional[int] = None,
        action: Optional[str] = None,
        entity_type: Optional[str] = None,
        entity_id: Optional[int] = None,
        entity_name: Optional[str] = None,
        search: Optional[str] = None,
        limit: int = 10000,
        event_name: Optional[str] = None
    ):
        """
        Exporta logs de auditoria para Excel (XLSX).

        **Vantagens do Excel:**
        - Formatação rica (cores, fontes, bordas)
        - Múltiplas abas (logs + estatísticas)
        - Fórmulas e cálculos automáticos
        - Filtros e ordenação nativos
        - Profissional para relatórios

        Retorna: (excel_content: bytes, filename: str, total: int)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            import io
        except ImportError:
            raise ImportError(
                "openpyxl não está instalado. "
                "Instale com: pip install openpyxl"
            )

        # Buscar logs com os filtros
        logs, total = self.search_logs_filtered(
            start_date=start_date,
            end_date=end_date,
            user_id=user_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            search=search,
            limit=limit,
            offset=0
        )

        # Criar workbook
        wb = Workbook()

        # Aba 1: Logs detalhados
        ws_logs = wb.active
        ws_logs.title = "Logs de Auditoria"

        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Cabeçalhos
        headers = [
            'ID', 'Data/Hora', 'Módulo', 'Entidade', 'ID Entidade',
            'Ação', 'Usuário', 'ID Usuário', 'Valores Antigos',
            'Valores Novos', 'Descrição'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws_logs.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Dados
        for row_num, log in enumerate(logs, 2):
            ws_logs.cell(row=row_num, column=1, value=log.id)
            ws_logs.cell(row=row_num, column=2, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws_logs.cell(row=row_num, column=3, value=getattr(log, 'entity_type', '').upper())
            ws_logs.cell(row=row_num, column=4, value=getattr(log, 'entity_name', ''))
            ws_logs.cell(row=row_num, column=5, value=getattr(log, 'entity_id', ''))
            ws_logs.cell(row=row_num, column=6, value=log.action.value if hasattr(log.action, 'value') else str(log.action))
            ws_logs.cell(row=row_num, column=7, value=getattr(log, 'created_by_username', ''))
            ws_logs.cell(row=row_num, column=8, value=log.created_by_id)
            ws_logs.cell(row=row_num, column=9, value=str(log.old_values) if log.old_values else '')
            ws_logs.cell(row=row_num, column=10, value=str(log.new_values) if log.new_values else '')
            ws_logs.cell(row=row_num, column=11, value=log.description or '')

        # Ajustar largura das colunas
        for col_num in range(1, len(headers) + 1):
            ws_logs.column_dimensions[get_column_letter(col_num)].width = 15

        # Aba 2: Estatísticas
        ws_stats = wb.create_sheet("Estatísticas")

        # Calcular estatísticas
        actions_count = {}
        entities_count = {}
        users_count = {}

        for log in logs:
            action = log.action.value if hasattr(log.action, 'value') else str(log.action)
            entity_type = getattr(log, 'entity_type', 'unknown')
            username = getattr(log, 'created_by_username', 'unknown')

            actions_count[action] = actions_count.get(action, 0) + 1
            entities_count[entity_type] = entities_count.get(entity_type, 0) + 1
            users_count[username] = users_count.get(username, 0) + 1

        # Estatísticas gerais
        ws_stats.cell(row=1, column=1, value="Resumo da Exportação").font = Font(bold=True, size=14)
        row = 2
        if event_name:
            ws_stats.cell(row=row, column=1, value="Evento:")
            ws_stats.cell(row=row, column=2, value=event_name).font = Font(bold=True)
            row += 1
        ws_stats.cell(row=row, column=1, value="Total de registros:")
        ws_stats.cell(row=row, column=2, value=total)
        row += 1
        ws_stats.cell(row=row, column=1, value="Exportado em:")
        ws_stats.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # Filtros aplicados
        row += 2
        ws_stats.cell(row=row, column=1, value="Filtros Aplicados").font = Font(bold=True, size=12)
        row += 1

        filters_added = False
        if start_date:
            ws_stats.cell(row=row, column=1, value="Data inicial:")
            ws_stats.cell(row=row, column=2, value=start_date.strftime('%Y-%m-%d %H:%M:%S'))
            row += 1
            filters_added = True
        if end_date:
            ws_stats.cell(row=row, column=1, value="Data final:")
            ws_stats.cell(row=row, column=2, value=end_date.strftime('%Y-%m-%d %H:%M:%S'))
            row += 1
            filters_added = True
        if user_id:
            ws_stats.cell(row=row, column=1, value="ID do usuário:")
            ws_stats.cell(row=row, column=2, value=user_id)
            row += 1
            filters_added = True
        if action:
            ws_stats.cell(row=row, column=1, value="Tipo de ação:")
            ws_stats.cell(row=row, column=2, value=action)
            row += 1
            filters_added = True
        if entity_type:
            ws_stats.cell(row=row, column=1, value="Módulo:")
            ws_stats.cell(row=row, column=2, value=entity_type.upper())
            row += 1
            filters_added = True
        if entity_id:
            ws_stats.cell(row=row, column=1, value="ID da entidade:")
            ws_stats.cell(row=row, column=2, value=entity_id)
            row += 1
            filters_added = True
        if entity_name:
            ws_stats.cell(row=row, column=1, value="Nome da entidade:")
            ws_stats.cell(row=row, column=2, value=entity_name)
            row += 1
            filters_added = True
        if search:
            ws_stats.cell(row=row, column=1, value="Busca por texto:")
            ws_stats.cell(row=row, column=2, value=search)
            row += 1
            filters_added = True

        if not filters_added:
            ws_stats.cell(row=row, column=1, value="Nenhum filtro aplicado")
            ws_stats.cell(row=row, column=2, value="Todos os registros")
            row += 1

        row += 1  # Espaço

        # Ações por tipo
        ws_stats.cell(row=row, column=1, value="Ações por Tipo").font = Font(bold=True)
        ws_stats.cell(row=row+1, column=1, value="Ação").font = Font(bold=True)
        ws_stats.cell(row=row+1, column=2, value="Quantidade").font = Font(bold=True)
        for idx, (action_name, count) in enumerate(sorted(actions_count.items()), start=row+2):
            ws_stats.cell(row=idx, column=1, value=action_name)
            ws_stats.cell(row=idx, column=2, value=count)

        # Entidades por tipo
        row = row + len(actions_count) + 4
        ws_stats.cell(row=row, column=1, value="Entidades por Tipo").font = Font(bold=True)
        ws_stats.cell(row=row+1, column=1, value="Módulo").font = Font(bold=True)
        ws_stats.cell(row=row+1, column=2, value="Quantidade").font = Font(bold=True)
        for idx, (entity, count) in enumerate(sorted(entities_count.items()), start=row+2):
            ws_stats.cell(row=idx, column=1, value=entity.upper())
            ws_stats.cell(row=idx, column=2, value=count)
            ws_stats.cell(row=idx, column=2, value=count)

        # Usuários mais ativos
        row = row + len(entities_count) + 4
        ws_stats.cell(row=row, column=1, value="Usuários Mais Ativos (Top 10)").font = Font(bold=True)
        ws_stats.cell(row=row+1, column=1, value="Usuário").font = Font(bold=True)
        ws_stats.cell(row=row+1, column=2, value="Ações").font = Font(bold=True)
        top_users = sorted(users_count.items(), key=lambda x: x[1], reverse=True)[:10]
        for idx, (user, count) in enumerate(top_users, start=row+2):
            ws_stats.cell(row=idx, column=1, value=user)
            ws_stats.cell(row=idx, column=2, value=count)

        # Ajustar largura
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 15

        # Salvar em bytes
        output = io.BytesIO()
        wb.save(output)
        excel_content = output.getvalue()
        output.close()

        # Gerar nome do arquivo
        filename = self._get_filename(
            'audit_logs', '.xlsx',
            start_date, end_date, user_id, action, entity_type, event_name
        )

        return excel_content, filename, total

    def export_logs_to_pdf(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        user_id: Optional[int] = None,
        action: Optional[str] = None,
        entity_type: Optional[str] = None,
        entity_id: Optional[int] = None,
        entity_name: Optional[str] = None,
        search: Optional[str] = None,
        limit: int = 1000,  # Limite menor para PDF (performance)
        event_name: Optional[str] = None
    ):
        """
        Exporta logs de auditoria para PDF.

        **Vantagens do PDF:**
        - Formato fixo (não editável)
        - Ideal para arquivamento e compliance
        - Profissional para apresentações
        - Imutável (auditoria)

        Retorna: (pdf_content: bytes, filename: str, total: int)
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            import io
        except ImportError:
            raise ImportError(
                "reportlab não está instalado. "
                "Instale com: pip install reportlab"
            )

        # Buscar logs com os filtros
        logs, total = self.search_logs_filtered(
            start_date=start_date,
            end_date=end_date,
            user_id=user_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            search=search,
            limit=limit,
            offset=0
        )

        # Criar PDF em memória
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )

        # Elementos do PDF
        elements = []

        # Título
        title = Paragraph("Relatório de Auditoria", title_style)
        elements.append(title)

        # Informações do relatório
        info_style = styles['Normal']
        if event_name:
            elements.append(Paragraph(f"<b>Evento:</b> {event_name}", info_style))
        elements.append(Paragraph(f"<b>Gerado em:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", info_style))
        elements.append(Paragraph(f"<b>Total de registros:</b> {total}", info_style))

        # Seção de filtros aplicados
        elements.append(Spacer(1, 0.3*cm))
        filter_style = ParagraphStyle('FilterStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#555555'))
        elements.append(Paragraph("<b>Filtros Aplicados:</b>", filter_style))

        filters_added = False
        if start_date or end_date:
            periodo = f"{start_date.strftime('%d/%m/%Y %H:%M') if start_date else 'início'} até {end_date.strftime('%d/%m/%Y %H:%M') if end_date else 'hoje'}"
            elements.append(Paragraph(f"• <b>Período:</b> {periodo}", filter_style))
            filters_added = True

        if user_id:
            elements.append(Paragraph(f"• <b>ID do usuário:</b> {user_id}", filter_style))
            filters_added = True

        if action:
            elements.append(Paragraph(f"• <b>Tipo de ação:</b> {action}", filter_style))
            filters_added = True

        if entity_type:
            elements.append(Paragraph(f"• <b>Módulo:</b> {entity_type.upper()}", filter_style))
            filters_added = True

        if entity_id:
            elements.append(Paragraph(f"• <b>ID da entidade:</b> {entity_id}", filter_style))
            filters_added = True

        if entity_name:
            elements.append(Paragraph(f"• <b>Nome da entidade:</b> {entity_name}", filter_style))
            filters_added = True

        if search:
            elements.append(Paragraph(f"• <b>Busca por texto:</b> {search}", filter_style))
            filters_added = True

        if not filters_added:
            elements.append(Paragraph("• Nenhum filtro aplicado (todos os registros)", filter_style))

        elements.append(Spacer(1, 0.5*cm))

        # Tabela de logs
        data = [['ID', 'Data/Hora', 'Módulo', 'Entidade', 'Ação', 'Usuário', 'Descrição']]

        for log in logs:
            data.append([
                str(log.id),
                log.created_at.strftime('%d/%m/%Y\n%H:%M'),
                getattr(log, 'entity_type', '').upper(),
                getattr(log, 'entity_name', '')[:20],  # Truncar para caber
                (log.action.value if hasattr(log.action, 'value') else str(log.action))[:15],
                getattr(log, 'created_by_username', '')[:15],
                (log.description or '')[:30]
            ])

        table = Table(data, colWidths=[1.5*cm, 3*cm, 2.5*cm, 4*cm, 3*cm, 3*cm, 6*cm])
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Dados
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(table)

        # Rodapé com disclaimer
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
        elements.append(Paragraph(
            "Este relatório foi gerado automaticamente pelo sistema de auditoria e contém informações confidenciais.",
            footer_style
        ))

        # Gerar PDF
        doc.build(elements)
        pdf_content = buffer.getvalue()
        buffer.close()

        # Gerar nome do arquivo
        filename = self._get_filename(
            'audit_logs', '.pdf',
            start_date, end_date, user_id, action, entity_type, event_name
        )

        return pdf_content, filename, total


# ============================================
# HELPER PARA DETECTAR MUDANÇAS
# ============================================

def get_changed_fields(old_obj, new_data: dict) -> tuple[dict, dict]:
    """
    Helper para detectar o que mudou em um objeto.

    Retorna: (old_values, new_values)

    Exemplo:
        old_vals, new_vals = get_changed_fields(customer, {"nome": "João Silva"})
        # old_vals = {"nome": "João"}
        # new_vals = {"nome": "João Silva"}
    """
    old_values = {}
    new_values = {}

    for field, new_value in new_data.items():
        if hasattr(old_obj, field):
            old_value = getattr(old_obj, field)

            # Só adicionar se realmente mudou
            if old_value != new_value:
                old_values[field] = old_value
                new_values[field] = new_value

    return old_values, new_values

