"""
Script para importar produtos da planilha Excel
Lê a aba CONSOLIDADO e importa todos os produtos com seus tipos
Rastreia lotes de importação para permitir rollback
"""
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import func
from pathlib import Path
from typing import Optional
from fastapi import UploadFile

from database import SessionLocal
from app.models import Produto, SystemUser, ProductType, ProductImportBatch, Restock, RestockBatch
from app.core.security import get_password_hash


def map_categoria_to_type(categoria: str) -> Optional[ProductType]:
    """
    Mapeia o nome da categoria da planilha para o enum ProductType

    Args:
        categoria: Nome da categoria (Doces, Salgados, Bebidas)

    Returns:
        ProductType correspondente ou None
    """
    if not categoria:
        return None

    categoria_lower = categoria.lower().strip()

    if categoria_lower in ['bebida', 'bebidas']:
        return ProductType.BEBIDA
    elif categoria_lower in ['doce', 'doces']:
        return ProductType.DOCE
    elif categoria_lower in ['salgadinho', 'salgadinhos', 'salgado', 'salgados']:
        return ProductType.SALGADINHO

    return None


def import_products_from_excel_file(
    file_path: str,
    db: Session,
    created_by_username: str = "admin"
) -> dict:
    """
    Importa produtos de um arquivo Excel

    Args:
        file_path: Caminho do arquivo Excel
        db: Sessão do banco de dados
        created_by_username: Username do usuário que está importando

    Returns:
        Dict com estatísticas da importação
    """

    # Buscar usuário que está importando
    user = db.query(SystemUser).filter(SystemUser.username == created_by_username).first()

    if not user:
        # Criar usuário admin se não existir
        from app.models import UserRole
        hashed_password = get_password_hash("admin123")
        user = SystemUser(
            username="admin",
            hashed_password=hashed_password,
            role=UserRole.ADMIN,
            is_active=True
        )
        db.add(user)
        db.commit()
        db.refresh(user)

    # Carregar planilha
    wb = load_workbook(file_path, data_only=True)

    # Verificar se existe aba CONSOLIDADO
    if 'CONSOLIDADO' not in wb.sheetnames:
        return {
            "success": False,
            "error": "Aba 'CONSOLIDADO' não encontrada na planilha",
            "imported": 0,
            "skipped": 0,
            "errors": 0
        }

    ws = wb['CONSOLIDADO']

    # Criar registro de batch de importação
    import_batch = ProductImportBatch(
        filename=Path(file_path).name,
        created_by_id=user.id,
        status="in_progress"
    )
    db.add(import_batch)
    db.commit()
    db.refresh(import_batch)

    # Contadores
    imported_count = 0
    skipped_count = 0
    error_count = 0
    errors_detail = []

    # Processar linhas (começar da linha 4, pular cabeçalhos)
    for row_num in range(4, ws.max_row + 1):
        # Ler dados da linha
        # Colunas da planilha:
        # 1. Categoria
        # 2. Nome do Produto
        # 3. Preço de compra (R$) - custo unitário ou do fardo
        # 4. Preço Unitário/Venda (R$)
        # 5. Nº de Fardos (opcional)
        # 6. Qtd por Fardo (opcional)
        # 7. Quantidade Total (calculada ou manual)
        # 8. Estoque Mínimo (opcional)

        categoria = ws.cell(row=row_num, column=1).value
        nome = ws.cell(row=row_num, column=2).value
        preco_compra = ws.cell(row=row_num, column=3).value  # Custo
        preco_venda = ws.cell(row=row_num, column=4).value   # Venda
        num_fardos = ws.cell(row=row_num, column=5).value
        qtd_fardo = ws.cell(row=row_num, column=6).value
        qtd_total = ws.cell(row=row_num, column=7).value
        estoque_minimo = ws.cell(row=row_num, column=8).value

        # Pular linhas vazias
        if not nome:
            continue

        # Validar se tem pelo menos preço de compra OU preço de venda
        if not preco_compra and not preco_venda:
            continue

        try:
            # Validar dados obrigatórios
            nome = str(nome).strip()

            # Converter preços para float (pode ser None)
            preco_compra_float = float(preco_compra) if preco_compra and preco_compra != '' else None
            preco_venda_float = float(preco_venda) if preco_venda and preco_venda != '' else None

            # Pelo menos um preço deve ser positivo
            if (preco_compra_float is None or preco_compra_float <= 0) and (preco_venda_float is None or preco_venda_float <= 0):
                skipped_count += 1
                errors_detail.append(f"Linha {row_num}: Preço inválido - produto '{nome}' ignorado")
                continue

            # Mapear categoria para tipo
            tipo = map_categoria_to_type(categoria) if categoria else None

            # ========================================
            # LÓGICA DE PRECIFICAÇÃO E ESTOQUE
            # ========================================

            # Verificar se tem informação de fardos
            tem_fardos = num_fardos and qtd_fardo and num_fardos != 0 and qtd_fardo != 0

            if tem_fardos:
                # ===== CENÁRIO COM FARDOS =====
                num_fardos_float = float(num_fardos)
                qtd_fardo_float = float(qtd_fardo)

                # Calcular estoque total = num_fardos × qtd_por_fardo
                estoque = int(num_fardos_float * qtd_fardo_float)

                # Preço de compra é o custo do FARDO
                # Dividir pela quantidade do fardo para obter custo unitário
                if preco_compra_float:
                    preco_custo_final = preco_compra_float / qtd_fardo_float
                else:
                    preco_custo_final = None

                # Preço de venda já é unitário
                preco_venda_final = preco_venda_float if preco_venda_float else None

                # Se não tem preço de venda, sugerir margem de 50%
                if preco_venda_final is None and preco_custo_final is not None:
                    preco_venda_final = preco_custo_final * 1.5

            else:
                # ===== CENÁRIO SEM FARDOS (UNIDADES) =====

                # Preço de compra e venda já são unitários
                preco_custo_final = preco_compra_float
                preco_venda_final = preco_venda_float

                # Estoque = quantidade total informada
                if qtd_total and qtd_total != '' and qtd_total != 0:
                    estoque = int(float(qtd_total))
                else:
                    estoque = 0

                # Se não tem preço de venda mas tem custo, sugerir margem
                if preco_venda_final is None and preco_custo_final is not None:
                    preco_venda_final = preco_custo_final * 1.5

            # Garantir que temos pelo menos um preço de venda
            if preco_venda_final is None:
                skipped_count += 1
                errors_detail.append(f"Linha {row_num}: Produto '{nome}' sem preço de venda - ignorado")
                continue

            # Estoque mínimo
            estoque_min = 10  # Padrão
            if estoque_minimo and estoque_minimo != '' and estoque_minimo != 0:
                estoque_min = int(float(estoque_minimo))

            # Verificar se produto já existe
            existing_product = db.query(Produto).filter(Produto.nome == nome).first()

            if existing_product:
                skipped_count += 1
                errors_detail.append(f"Linha {row_num}: Produto '{nome}' já existe - ignorado")
                continue

            # Criar produto
            produto = Produto(
                nome=nome,
                tipo=tipo,
                valor=preco_venda_final,  # Compatibilidade
                preco_custo=preco_custo_final,
                preco_venda=preco_venda_final,
                estoque=estoque,
                estoque_minimo=estoque_min,
                is_active=True,
                created_by_id=user.id,
                import_batch_id=import_batch.id  # Associar ao batch
            )

            db.add(produto)
            db.commit()
            db.refresh(produto)

            imported_count += 1

        except Exception as e:
            error_count += 1
            errors_detail.append(f"Linha {row_num}: Erro ao importar produto '{nome}': {str(e)}")
            db.rollback()

    # Atualizar estatísticas do batch
    import_batch.imported_count = imported_count
    import_batch.skipped_count = skipped_count
    import_batch.error_count = error_count
    import_batch.status = "completed"
    db.commit()

    # 🆕 AUDITORIA: Registrar importação
    from app.services.audit import AuditService
    from app.models_audit import AuditAction

    audit = AuditService(db)
    audit.log_system_action(
        action=AuditAction.IMPORT,
        created_by_id=user.id,
        entity_type="product_batch",
        entity_id=import_batch.id,
        new_values={
            "filename": Path(file_path).name,
            "imported_count": imported_count,
            "skipped_count": skipped_count,
            "error_count": error_count,
            "batch_id": import_batch.id
        },
        description=f"Importação em massa: {imported_count} produto(s) importado(s) de '{Path(file_path).name}' por {user.username}"
    )

    return {
        "success": True,
        "batch_id": import_batch.id,
        "imported": imported_count,
        "skipped": skipped_count,
        "errors": error_count,
        "errors_detail": errors_detail[:10]  # Limitar a 10 erros
    }


async def import_products_from_upload(
    upload_file: UploadFile,
    db: Session,
    created_by_username: str = "admin"
) -> dict:
    """
    Importa produtos de um arquivo Excel enviado via upload

    Args:
        upload_file: Arquivo Excel enviado
        db: Sessão do banco de dados
        created_by_username: Username do usuário que está importando

    Returns:
        Dict com estatísticas da importação
    """
    import tempfile

    # Salvar arquivo temporariamente
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        content = await upload_file.read()
        tmp_file.write(content)
        tmp_path = tmp_file.name

    try:
        # Importar do arquivo temporário
        result = import_products_from_excel_file(tmp_path, db, created_by_username)
        return result
    finally:
        # Limpar arquivo temporário
        Path(tmp_path).unlink(missing_ok=True)


def rollback_import_batch(
    batch_id: int,
    db: Session,
    rolled_back_by_username: str
) -> dict:
    """
    Faz rollback de uma importação, deletando todos os produtos importados naquele batch

    Args:
        batch_id: ID do batch de importação
        db: Sessão do banco de dados
        rolled_back_by_username: Username do usuário fazendo o rollback

    Returns:
        Dict com resultado do rollback
    """
    from datetime import datetime, timezone
    from app.core.timezone import get_now

    # Buscar batch
    import_batch = db.query(ProductImportBatch).filter(ProductImportBatch.id == batch_id).first()

    if not import_batch:
        return {
            "success": False,
            "error": "Batch de importação não encontrado"
        }

    # Verificar se já foi revertido
    if import_batch.status == "rolled_back":
        return {
            "success": False,
            "error": "Este batch já foi revertido anteriormente"
        }

    # Buscar usuário
    user = db.query(SystemUser).filter(SystemUser.username == rolled_back_by_username).first()

    if not user:
        return {
            "success": False,
            "error": "Usuário não encontrado"
        }

    # Buscar todos os produtos deste batch
    products = db.query(Produto).filter(Produto.import_batch_id == batch_id).all()

    if not products:
        return {
            "success": False,
            "error": "Nenhum produto encontrado para este batch"
        }

    # Verificar se algum produto foi vendido
    from app.models import SaleItem
    products_with_sales = []
    for product in products:
        sales_count = db.query(SaleItem).filter(SaleItem.produto_id == product.id).count()
        if sales_count > 0:
            products_with_sales.append({
                "id": product.id,
                "nome": product.nome,
                "sales_count": sales_count
            })

    if products_with_sales:
        return {
            "success": False,
            "error": "Não é possível reverter: alguns produtos já foram vendidos",
            "products_with_sales": products_with_sales[:5]  # Mostrar até 5 produtos
        }

    # Deletar produtos
    deleted_count = 0
    deleted_products = []
    for product in products:
        deleted_products.append({
            "id": product.id,
            "nome": product.nome,
            "tipo": product.tipo.value if product.tipo else None,
            "valor": product.valor,
            "estoque": product.estoque
        })
        db.delete(product)
        deleted_count += 1

    # Atualizar batch
    import_batch.status = "rolled_back"
    import_batch.rolled_back_at = get_now()
    import_batch.rolled_back_by_id = user.id

    db.commit()

    # 🆕 AUDITORIA: Registrar rollback
    from app.services.audit import AuditService
    from app.models_audit import AuditAction

    audit = AuditService(db)
    audit.log_system_action(
        action=AuditAction.ROLLBACK,
        created_by_id=user.id,
        entity_type="product_batch",
        entity_id=batch_id,
        old_values={
            "status": "completed",
            "products_count": deleted_count,
            "products": deleted_products[:10]  # Limitar a 10 produtos no log
        },
        new_values={
            "status": "rolled_back"
        },
        description=f"Rollback de importação: {deleted_count} produto(s) deletado(s) do batch #{batch_id} (arquivo: {import_batch.filename}) por {user.username}"
    )

    return {
        "success": True,
        "batch_id": batch_id,
        "deleted_count": deleted_count,
        "rolled_back_by": rolled_back_by_username,
        "rolled_back_at": import_batch.rolled_back_at.isoformat()
    }


def get_import_batches_list(
    db: Session,
    skip: int = 0,
    limit: int = 50,
    include_rolled_back: bool = True
) -> list:
    """
    Lista todos os batches de importação

    Args:
        db: Sessão do banco de dados
        skip: Número de registros para pular
        limit: Número máximo de registros
        include_rolled_back: Se deve incluir batches revertidos

    Returns:
        Lista de batches de importação
    """
    query = db.query(ProductImportBatch)

    if not include_rolled_back:
        query = query.filter(ProductImportBatch.status != "rolled_back")

    batches = query.order_by(ProductImportBatch.created_at.desc()).offset(skip).limit(limit).all()

    result = []
    for batch in batches:
        # Contar produtos restantes (não deletados)
        products_count = db.query(Produto).filter(Produto.import_batch_id == batch.id).count()

        result.append({
            "id": batch.id,
            "filename": batch.filename,
            "imported_count": batch.imported_count,
            "skipped_count": batch.skipped_count,
            "error_count": batch.error_count,
            "current_products_count": products_count,
            "status": batch.status,
            "created_at": batch.created_at.isoformat(),
            "created_by": batch.created_by.username if batch.created_by else None,
            "rolled_back_at": batch.rolled_back_at.isoformat() if batch.rolled_back_at else None,
            "rolled_back_by": batch.rolled_back_by.username if batch.rolled_back_by else None,
            "can_rollback": batch.status == "completed" and products_count > 0
        })

    return result


# ============================================
# REABASTECIMENTO EM MASSA VIA EXCEL
# ============================================

def import_bulk_restock_from_excel_file(
    file_path: str,
    db: Session,
    created_by_username: str = "admin"
) -> dict:
    """
    Importa reabastecimentos em massa de um arquivo Excel

    Args:
        file_path: Caminho do arquivo Excel
        db: Sessão do banco de dados
        created_by_username: Username do usuário que está importando

    Returns:
        Dict com estatísticas e resultados da importação
    """
    # Buscar usuário que está importando
    user = db.query(SystemUser).filter(SystemUser.username == created_by_username).first()

    if not user:
        return {
            "success": False,
            "error": "Usuário não encontrado",
            "processed": 0,
            "succeeded": 0,
            "failed": 0,
            "not_found": 0,
            "results": []
        }

    # Carregar planilha
    wb = load_workbook(file_path, data_only=True)

    # Verificar se existe aba REABASTECIMENTO
    if 'REABASTECIMENTO' not in wb.sheetnames:
        return {
            "success": False,
            "error": "Aba 'REABASTECIMENTO' não encontrada na planilha",
            "processed": 0,
            "succeeded": 0,
            "failed": 0,
            "not_found": 0,
            "results": []
        }

    ws = wb['REABASTECIMENTO']

    # Extrair nome do arquivo
    from pathlib import Path
    filename = Path(file_path).name

    # Criar batch para rastrear esta importação
    batch = RestockBatch(
        filename=filename,
        created_by_id=user.id
    )
    db.add(batch)
    db.flush()  # Para obter o ID do batch

    # Contadores
    processed_count = 0
    success_count = 0
    error_count = 0
    not_found_count = 0
    results = []
    not_found_products = []

    # ============================================
    # DETECTAR LINHA INICIAL DOS DADOS
    # ============================================
    # Procurar pela linha do cabeçalho para determinar onde começam os dados
    header_row = None
    data_start_row = None

    # Procurar cabeçalho nas primeiras 10 linhas
    for row_num in range(1, min(11, ws.max_row + 1)):
        cell_a = ws.cell(row=row_num, column=1).value
        cell_b = ws.cell(row=row_num, column=2).value

        # Se encontrar texto que parece cabeçalho
        if cell_a and isinstance(cell_a, str):
            cell_a_str = str(cell_a).lower().strip()
            if "código" in cell_a_str or "codigo" in cell_a_str:
                header_row = row_num
                data_start_row = row_num + 1
                break

        if cell_b and isinstance(cell_b, str):
            cell_b_str = str(cell_b).lower().strip()
            if "nome" in cell_b_str and "produto" in cell_b_str:
                header_row = row_num
                data_start_row = row_num + 1
                break

    # Se não encontrou cabeçalho, assumir que dados começam na linha 2
    if data_start_row is None:
        data_start_row = 2

    # Limitar até linha 100 (segurança)
    max_data_row = min(data_start_row + 100, ws.max_row + 1)

    # Processar linhas de dados
    for row_num in range(data_start_row, max_data_row):
        # Ler dados da linha
        # Colunas:
        # A: Código do Produto (opcional)
        # B: Nome do Produto (opcional)
        # C: Nº de Fardos (opcional)
        # D: Qtd por Fardo (opcional)
        # E: Quantidade Total (pode ser calculada ou manual)

        codigo_cell = ws.cell(row=row_num, column=1).value
        nome_cell = ws.cell(row=row_num, column=2).value
        num_fardos_cell = ws.cell(row=row_num, column=3).value
        qtd_fardo_cell = ws.cell(row=row_num, column=4).value
        quantidade_total_cell = ws.cell(row=row_num, column=5).value

        # Debug da linha
        print(f"📊 Linha {row_num}: cod={codigo_cell}, nome={nome_cell}, fardos={num_fardos_cell}, qtd_fardo={qtd_fardo_cell}, total={quantidade_total_cell}")

        # Pular linhas vazias (nem código, nem nome, nem quantidade)
        if not codigo_cell and not nome_cell and not quantidade_total_cell and not num_fardos_cell:
            print(f"⏭️ Linha {row_num}: VAZIA - pulando")
            continue

        # Ignorar linhas de cabeçalho duplicadas
        if codigo_cell and isinstance(codigo_cell, str):
            codigo_str = str(codigo_cell).lower().strip()
            # Se contém palavras de cabeçalho, pular
            if any(palavra in codigo_str for palavra in ["código", "codigo", "nome do produto", "cabeçalho"]):
                print(f"⏭️ Linha {row_num}: CABEÇALHO duplicado detectado - pulando")
                continue

        if nome_cell and isinstance(nome_cell, str):
            nome_str = str(nome_cell).lower().strip()
            # Se contém palavras de cabeçalho, pular
            if any(palavra in nome_str for palavra in ["nome do produto", "nome produto", "produto"]):
                print(f"⏭️ Linha {row_num}: CABEÇALHO duplicado detectado - pulando")
                continue

        # Ignorar linhas que são claramente instruções
        if codigo_cell and isinstance(codigo_cell, str):
            codigo_str = str(codigo_cell).strip()
            # 1. Emojis (caracteres especiais)
            # 2. Texto muito longo (>25 caracteres)
            # 3. Contém ":" ou "INSTRUÇÃO"
            if (any(ord(char) > 127 for char in codigo_str) or  # Emoji/Unicode
                len(codigo_str) > 25 or
                ":" in codigo_str or
                "instrução" in codigo_str.lower() or
                "instruções" in codigo_str.lower()):
                print(f"⏭️ Linha {row_num}: INSTRUÇÃO detectada - pulando")
                continue

        # Validar se tem pelo menos código OU nome
        if not codigo_cell and not nome_cell:
            results.append({
                "row": row_num,
                "codigo": None,
                "nome": None,
                "quantidade": quantidade_total_cell,
                "status": "error",
                "message": "Linha ignorada: nem código nem nome fornecido",
                "produto_id": None,
                "produto_nome_encontrado": None,
                "estoque_anterior": None,
                "estoque_atual": None
            })
            error_count += 1
            processed_count += 1
            continue

        # Calcular quantidade total
        # Se tem fardos, calcular: num_fardos × qtd_por_fardo
        # Senão, usar quantidade_total diretamente
        quantidade = 0
        tem_fardos = num_fardos_cell and qtd_fardo_cell

        try:
            if tem_fardos:
                # COM FARDOS: calcular quantidade
                num_fardos = float(num_fardos_cell)
                qtd_fardo = float(qtd_fardo_cell)

                if num_fardos <= 0 or qtd_fardo <= 0:
                    raise ValueError("Número de fardos e quantidade por fardo devem ser positivos")

                quantidade = int(num_fardos * qtd_fardo)
            else:
                # SEM FARDOS: usar quantidade total diretamente
                if not quantidade_total_cell:
                    raise ValueError("Quantidade total não informada")

                quantidade = int(float(quantidade_total_cell))

            # Validar quantidade final
            if quantidade <= 0:
                raise ValueError("Quantidade deve ser maior que zero")

        except (ValueError, TypeError) as e:
            results.append({
                "row": row_num,
                "codigo": str(codigo_cell) if codigo_cell else None,
                "nome": str(nome_cell) if nome_cell else None,
                "quantidade": quantidade_total_cell,
                "status": "error",
                "message": f"Quantidade inválida: {str(e)}",
                "produto_id": None,
                "produto_nome_encontrado": None,
                "estoque_anterior": None,
                "estoque_atual": None
            })
            error_count += 1
            processed_count += 1
            continue

        # Buscar produto por código OU nome (case insensitive)
        produto = None

        # Tentar por código primeiro (se fornecido)
        if codigo_cell:
            try:
                produto_id = int(codigo_cell)
                produto = db.query(Produto).filter(Produto.id == produto_id).first()
            except (ValueError, TypeError):
                pass  # Código não é numérico, tentar por nome

        # Se não encontrou por código, tentar por nome (case insensitive)
        if not produto and nome_cell:
            nome_busca = str(nome_cell).strip()
            produto = db.query(Produto).filter(
                func.lower(Produto.nome) == func.lower(nome_busca)
            ).first()

        # Se não encontrou o produto
        if not produto:
            results.append({
                "row": row_num,
                "codigo": str(codigo_cell) if codigo_cell else None,
                "nome": str(nome_cell) if nome_cell else None,
                "quantidade": quantidade,
                "status": "not_found",
                "message": "Produto não encontrado",
                "produto_id": None,
                "produto_nome_encontrado": None,
                "estoque_anterior": None,
                "estoque_atual": None
            })
            not_found_products.append({
                "row": row_num,
                "codigo": str(codigo_cell) if codigo_cell else None,
                "nome": str(nome_cell) if nome_cell else None
            })
            not_found_count += 1
            processed_count += 1
            continue

        # Produto encontrado - fazer reabastecimento
        try:
            estoque_anterior = produto.estoque
            produto.estoque += quantidade

            # Criar registro de reabastecimento ASSOCIADO AO BATCH
            restock = Restock(
                produto_id=produto.id,
                created_by_id=user.id,
                quantity=quantidade,
                batch_id=batch.id  # ← Associar ao batch
            )
            db.add(restock)
            db.flush()  # ← Usar flush ao invés de commit (commit só no final)

            # 🆕 AUDITORIA: Registrar reabastecimento individual
            from app.services.audit import AuditService
            from app.models_audit import AuditAction

            audit = AuditService(db)
            audit.log_product_action(
                produto_id=produto.id,
                action=AuditAction.RESTOCK,
                created_by_id=user.id,
                old_values={"estoque": estoque_anterior},
                new_values={
                    "estoque": produto.estoque,
                    "quantidade_adicionada": quantidade,
                    "batch_id": batch.id,
                    "restock_id": restock.id
                },
                description=f"Reabastecimento via importação em massa: +{quantidade} unidade(s) - Estoque: {estoque_anterior} → {produto.estoque} (Batch #{batch.id})"
            )

            results.append({
                "row": row_num,
                "codigo": str(codigo_cell) if codigo_cell else None,
                "nome": str(nome_cell) if nome_cell else None,
                "quantidade": quantidade,
                "status": "success",
                "message": f"Reabastecido com sucesso",
                "produto_id": produto.id,
                "produto_nome_encontrado": produto.nome,
                "estoque_anterior": estoque_anterior,
                "estoque_atual": produto.estoque
            })
            success_count += 1
            processed_count += 1

        except Exception as e:
            # NÃO fazer rollback aqui, apenas registrar o erro
            # Rollback desfaria todo o batch incluindo sucessos anteriores
            results.append({
                "row": row_num,
                "codigo": str(codigo_cell) if codigo_cell else None,
                "nome": str(nome_cell) if nome_cell else None,
                "quantidade": quantidade,
                "status": "error",
                "message": f"Erro ao reabastecer: {str(e)}",
                "produto_id": produto.id if produto else None,
                "produto_nome_encontrado": produto.nome if produto else None,
                "estoque_anterior": None,
                "estoque_atual": None
            })
            error_count += 1
            processed_count += 1
            # Continuar para próxima linha mesmo com erro

    # Atualizar estatísticas do batch
    batch.succeeded_count = success_count
    batch.failed_count = error_count
    batch.not_found_count = not_found_count
    batch.status = "completed"
    db.commit()

    # 🆕 AUDITORIA: Registrar importação em massa (IMPORT, não RESTOCK)
    from app.services.audit import AuditService
    from app.models_audit import AuditAction

    audit = AuditService(db)
    audit.log_system_action(
        action=AuditAction.IMPORT,
        created_by_id=user.id,
        entity_type="restock_batch",
        entity_id=batch.id,
        new_values={
            "filename": filename,
            "succeeded_count": success_count,
            "failed_count": error_count,
            "not_found_count": not_found_count,
            "batch_id": batch.id
        },
        description=f"Importação de reabastecimento: {success_count} produto(s) reabastecido(s) de '{filename}' por {user.username}"
    )

    return {
        "success": True,
        "batch_id": batch.id,
        "total_rows": processed_count,
        "processed": processed_count,
        "succeeded": success_count,
        "failed": error_count,
        "not_found": not_found_count,
        "results": results,
        "not_found_products": not_found_products
    }


async def import_bulk_restock_from_upload(
    upload_file: UploadFile,
    db: Session,
    created_by_username: str = "admin"
) -> dict:
    """
    Importa reabastecimentos de um arquivo Excel enviado via upload

    Args:
        upload_file: Arquivo Excel enviado
        db: Sessão do banco de dados
        created_by_username: Username do usuário que está importando

    Returns:
        Dict com estatísticas da importação
    """
    import tempfile

    # Salvar arquivo temporariamente
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        content = await upload_file.read()
        tmp_file.write(content)
        tmp_path = tmp_file.name

    try:
        # Importar do arquivo temporário
        result = import_bulk_restock_from_excel_file(tmp_path, db, created_by_username)
        return result
    finally:
        # Limpar arquivo temporário
        Path(tmp_path).unlink(missing_ok=True)


# ============================================
# ROLLBACK DE REABASTECIMENTO EM MASSA
# ============================================

def rollback_restock_batch(
    batch_id: int,
    db: Session,
    rolled_back_by_username: str
) -> dict:
    """
    Faz rollback de uma importação de reabastecimento, revertendo todos os restocks do batch

    Args:
        batch_id: ID do batch de reabastecimento
        db: Sessão do banco de dados
        rolled_back_by_username: Username do usuário fazendo o rollback

    Returns:
        Dict com resultado do rollback
    """
    from app.core.timezone import get_now

    # Buscar batch
    batch = db.query(RestockBatch).filter(RestockBatch.id == batch_id).first()

    if not batch:
        return {
            "success": False,
            "error": "Batch de reabastecimento não encontrado"
        }

    # Verificar se já foi revertido
    if batch.status == "rolled_back":
        return {
            "success": False,
            "error": "Este batch já foi revertido anteriormente"
        }

    # Buscar usuário
    user = db.query(SystemUser).filter(SystemUser.username == rolled_back_by_username).first()

    if not user:
        return {
            "success": False,
            "error": "Usuário não encontrado"
        }

    # Buscar todos os restocks deste batch
    restocks = db.query(Restock).filter(Restock.batch_id == batch_id).all()

    if not restocks:
        return {
            "success": False,
            "error": "Nenhum reabastecimento encontrado para este batch"
        }

    # Reverter cada restock (subtrair quantidade do estoque)
    reverted_count = 0
    reverted_restocks = []

    try:
        for restock in restocks:
            produto = db.query(Produto).filter(Produto.id == restock.produto_id).first()

            if not produto:
                continue

            # Verificar se há estoque suficiente para reverter
            if produto.estoque < restock.quantity:
                return {
                    "success": False,
                    "error": f"Não é possível reverter: produto '{produto.nome}' tem estoque insuficiente (atual: {produto.estoque}, necessário: {restock.quantity})"
                }

            # Guardar valores antes da reversão
            estoque_antes = produto.estoque

            # Subtrair quantidade do estoque
            produto.estoque -= restock.quantity

            reverted_restocks.append({
                "id": restock.id,
                "produto_id": produto.id,
                "produto_nome": produto.nome,
                "quantidade": restock.quantity,
                "estoque_antes_rollback": estoque_antes,
                "estoque_depois_rollback": produto.estoque
            })

            # 🆕 AUDITORIA: Registrar reversão individual
            from app.services.audit import AuditService
            from app.models_audit import AuditAction

            audit = AuditService(db)
            audit.log_product_action(
                produto_id=produto.id,
                action=AuditAction.ROLLBACK,
                created_by_id=user.id,
                old_values={
                    "estoque": estoque_antes,
                    "restock_id": restock.id,
                    "batch_id": batch_id
                },
                new_values={
                    "estoque": produto.estoque,
                    "quantidade_removida": restock.quantity
                },
                description=f"Rollback de reabastecimento: -{restock.quantity} unidade(s) - Estoque: {estoque_antes} → {produto.estoque} (Batch #{batch_id} revertido)"
            )

            # Deletar registro de restock
            db.delete(restock)
            reverted_count += 1

        # Atualizar batch
        batch.status = "rolled_back"
        batch.rolled_back_at = get_now()
        batch.rolled_back_by_id = user.id

        db.commit()

        # 🆕 AUDITORIA: Registrar rollback do batch completo
        from app.services.audit import AuditService
        from app.models_audit import AuditAction

        audit = AuditService(db)
        audit.log_system_action(
            action=AuditAction.ROLLBACK,
            created_by_id=user.id,
            entity_type="restock_batch",
            entity_id=batch_id,
            old_values={
                "status": "completed",
                "restocks_count": reverted_count,
                "restocks": reverted_restocks[:10]  # Limitar a 10 no log
            },
            new_values={
                "status": "rolled_back"
            },
            description=f"Rollback de reabastecimento: {reverted_count} reabastecimento(s) revertido(s) do batch #{batch_id} (arquivo: {batch.filename}) por {user.username}"
        )

        return {
            "success": True,
            "batch_id": batch_id,
            "reverted_count": reverted_count,
            "rolled_back_by": rolled_back_by_username,
            "rolled_back_at": batch.rolled_back_at.isoformat(),
            "reverted_restocks": reverted_restocks
        }

    except Exception as e:
        db.rollback()
        return {
            "success": False,
            "error": f"Erro ao fazer rollback: {str(e)}"
        }


def get_restock_batches_list(
    db: Session,
    skip: int = 0,
    limit: int = 50,
    include_rolled_back: bool = True
) -> list:
    """
    Lista todos os batches de reabastecimento

    Args:
        db: Sessão do banco de dados
        skip: Número de registros para pular
        limit: Número máximo de registros
        include_rolled_back: Se deve incluir batches revertidos

    Returns:
        Lista de batches de reabastecimento
    """
    query = db.query(RestockBatch)

    if not include_rolled_back:
        query = query.filter(RestockBatch.status != "rolled_back")

    batches = query.order_by(RestockBatch.created_at.desc()).offset(skip).limit(limit).all()

    result = []
    for batch in batches:
        # Contar restocks restantes (não deletados)
        restocks_count = db.query(Restock).filter(Restock.batch_id == batch.id).count()

        result.append({
            "id": batch.id,
            "filename": batch.filename,
            "succeeded_count": batch.succeeded_count,
            "failed_count": batch.failed_count,
            "not_found_count": batch.not_found_count,
            "current_restocks_count": restocks_count,
            "status": batch.status,
            "created_at": batch.created_at.isoformat(),
            "created_by": batch.created_by.username if batch.created_by else None,
            "rolled_back_at": batch.rolled_back_at.isoformat() if batch.rolled_back_at else None,
            "rolled_back_by": batch.rolled_back_by.username if batch.rolled_back_by else None,
            "can_rollback": batch.status == "completed" and restocks_count > 0
        })

    return result


# ============================================
# IMPORTAÇÃO DE CLIENTES VIA EXCEL
# ============================================

def import_customers_from_excel_file(
    file_path: str,
    db: Session,
    created_by_username: str = "admin"
) -> dict:
    """
    Importa clientes de um arquivo Excel com duas abas: ACAMPANTES e EQUIPE

    Args:
        file_path: Caminho do arquivo Excel
        db: Sessão do banco de dados
        created_by_username: Username do usuário que está importando

    Returns:
        Dict com estatísticas da importação
    """
    from app.models import Customers, CustomerTipo, CustomerImportBatch, EventConfig, EventRoom

    # Buscar usuário que está importando
    user = db.query(SystemUser).filter(SystemUser.username == created_by_username).first()

    if not user:
        return {
            "success": False,
            "error": "Usuário não encontrado",
            "imported": 0,
            "skipped": 0,
            "errors": 0
        }

    # Carregar planilha
    wb = load_workbook(file_path, data_only=True)

    # Verificar se existem as abas necessárias
    if 'ACAMPANTES' not in wb.sheetnames and 'EQUIPE' not in wb.sheetnames:
        return {
            "success": False,
            "error": "Nenhuma aba válida encontrada. Abas esperadas: 'ACAMPANTES' ou 'EQUIPE'",
            "imported": 0,
            "skipped": 0,
            "errors": 0
        }

    # Extrair nome do arquivo
    from pathlib import Path
    filename = Path(file_path).name

    # Criar batch para rastrear esta importação
    batch = CustomerImportBatch(
        filename=filename,
        created_by_id=user.id
    )
    db.add(batch)
    db.flush()  # Para obter o ID do batch

    # Contadores
    imported_count = 0
    skipped_count = 0
    error_count = 0
    errors_detail = []

    # Buscar quarto padrão "N/A" ou criar se não existir
    default_room = db.query(EventRoom).filter(EventRoom.room_name == "N/A").first()
    if not default_room:
        # Buscar configuração de evento ativa
        active_config = db.query(EventConfig).filter(EventConfig.is_active == True).first()
        if active_config:
            default_room = EventRoom(
                room_name="N/A",
                event_config_id=active_config.id,
                created_by_id=user.id
            )
            db.add(default_room)
            db.flush()

    # ============================================
    # PROCESSAR ABA: ACAMPANTES
    # ============================================
    if 'ACAMPANTES' in wb.sheetnames:
        ws = wb['ACAMPANTES']

        # Detectar linha do cabeçalho
        header_row = None
        for row_num in range(1, min(20, ws.max_row + 1)):
            cell_a = ws.cell(row=row_num, column=1).value
            if cell_a and isinstance(cell_a, str):
                cell_a_str = str(cell_a).strip()
                # Procurar por "Nome Completo" no cabeçalho
                if "Nome Completo" in cell_a_str and "*" in cell_a_str:
                    header_row = row_num
                    break

        # Se não encontrou cabeçalho, pular esta aba
        if header_row is None:
            errors_detail.append("ACAMPANTES: Cabeçalho não encontrado")
        else:
            # Dados começam na linha após o cabeçalho
            data_start = header_row + 1

            # Processar linhas de dados (limitar a 100 linhas após cabeçalho)
            for row_num in range(data_start, min(data_start + 100, ws.max_row + 1)):
                # Ler dados da linha
                # Colunas:
                # A: Nome Completo *
                # B: Apelido *
                # C: Quarto
                # D: Nome do Pai
                # E: Nome da Mãe
                # F: Saldo Inicial

                nome = ws.cell(row=row_num, column=1).value
                nickname = ws.cell(row=row_num, column=2).value
                quarto = ws.cell(row=row_num, column=3).value
                nome_pai = ws.cell(row=row_num, column=4).value
                nome_mae = ws.cell(row=row_num, column=5).value
                saldo = ws.cell(row=row_num, column=6).value

                # Pular linhas vazias (nem nome nem nickname)
                if not nome and not nickname:
                    continue

                # Ignorar linhas que contêm emojis ou caracteres especiais (instruções)
                if nome and isinstance(nome, str):
                    nome_str = str(nome).strip()
                    # Se contém emoji ou caracteres Unicode especiais, pular
                    if any(ord(char) > 127 for char in nome_str):
                        continue
                    # Se é muito longo (>100 chars), provavelmente é instrução
                    if len(nome_str) > 100:
                        continue

                # Validar campos obrigatórios
                if not nome or not nickname:
                    errors_detail.append(f"Linha {row_num} (ACAMPANTES): Nome Completo e Apelido são obrigatórios")
                    error_count += 1
                    continue

                try:
                    # Validar e processar dados
                    nome = str(nome).strip()
                    nickname = str(nickname).strip()

                    # Verificar se cliente já existe (nome + nickname)
                    existing = db.query(Customers).filter(
                        Customers.nome == nome,
                        Customers.nickname == nickname
                    ).first()

                    if existing:
                        skipped_count += 1
                        errors_detail.append(f"Linha {row_num} (ACAMPANTES): Cliente '{nome}' ({nickname}) já existe - ignorado")
                        continue

                    # Processar quarto
                    quarto_final = "N/A"  # Padrão
                    if quarto and str(quarto).strip():
                        quarto_name = str(quarto).strip()
                        # Verificar se o quarto existe no EventConfig
                        room = db.query(EventRoom).filter(
                            func.lower(EventRoom.room_name) == func.lower(quarto_name)
                        ).first()
                        if room:
                            quarto_final = room.room_name
                        else:
                            quarto_final = quarto_name  # Usar o nome fornecido mesmo que não exista

                    # Processar saldo
                    balance = 0.0
                    if saldo:
                        try:
                            balance = float(saldo)
                        except (ValueError, TypeError):
                            balance = 0.0

                    # Criar cliente
                    customer = Customers(
                        nome=nome,
                        nickname=nickname,
                        tipo=CustomerTipo.ACAMPANTE,
                        quarto=quarto_final,
                        nome_pai=str(nome_pai).strip() if nome_pai else None,
                        nome_mae=str(nome_mae).strip() if nome_mae else None,
                        saldo=balance,
                        import_batch_id=batch.id,
                        created_by_id=user.id
                    )

                    db.add(customer)
                    db.flush()

                    imported_count += 1

                except Exception as e:
                    error_count += 1
                    errors_detail.append(f"Linha {row_num} (ACAMPANTES): Erro ao importar '{nome}': {str(e)}")
                    continue

    # ============================================
    # PROCESSAR ABA: EQUIPE
    # ============================================
    if 'EQUIPE' in wb.sheetnames:
        ws = wb['EQUIPE']

        # Detectar linha do cabeçalho
        header_row = None
        for row_num in range(1, min(20, ws.max_row + 1)):
            cell_a = ws.cell(row=row_num, column=1).value
            if cell_a and isinstance(cell_a, str):
                cell_a_str = str(cell_a).strip()
                # Procurar por "Nome Completo" no cabeçalho
                if "Nome Completo" in cell_a_str and "*" in cell_a_str:
                    header_row = row_num
                    break

        # Se não encontrou cabeçalho, pular esta aba
        if header_row is None:
            errors_detail.append("EQUIPE: Cabeçalho não encontrado")
        else:
            # Dados começam na linha após o cabeçalho
            data_start = header_row + 1

            # Processar linhas de dados (limitar a 100 linhas após cabeçalho)
            for row_num in range(data_start, min(data_start + 100, ws.max_row + 1)):
                # Ler dados da linha
                # Colunas:
                # A: Nome Completo *
                # B: Apelido *
                # C: Nome do Pai
                # D: Nome da Mãe
                # E: Saldo Inicial

                nome = ws.cell(row=row_num, column=1).value
                nickname = ws.cell(row=row_num, column=2).value
                nome_pai = ws.cell(row=row_num, column=3).value
                nome_mae = ws.cell(row=row_num, column=4).value
                saldo = ws.cell(row=row_num, column=5).value

                # Pular linhas vazias
                if not nome and not nickname:
                    continue

                # Ignorar linhas que contêm emojis ou caracteres especiais (instruções)
                if nome and isinstance(nome, str):
                    nome_str = str(nome).strip()
                    # Se contém emoji ou caracteres Unicode especiais, pular
                    if any(ord(char) > 127 for char in nome_str):
                        continue
                    # Se é muito longo (>100 chars), provavelmente é instrução
                    if len(nome_str) > 100:
                        continue

                # Validar campos obrigatórios
                if not nome or not nickname:
                    errors_detail.append(f"Linha {row_num} (EQUIPE): Nome Completo e Apelido são obrigatórios")
                    error_count += 1
                    continue

                try:
                    # Validar e processar dados
                    nome = str(nome).strip()
                    nickname = str(nickname).strip()

                    # Verificar se cliente já existe
                    existing = db.query(Customers).filter(
                        Customers.nome == nome,
                        Customers.nickname == nickname
                    ).first()

                    if existing:
                        skipped_count += 1
                        errors_detail.append(f"Linha {row_num} (EQUIPE): Cliente '{nome}' ({nickname}) já existe - ignorado")
                        continue

                    # Processar saldo
                    balance = 0.0
                    if saldo:
                        try:
                            balance = float(saldo)
                        except (ValueError, TypeError):
                            balance = 0.0

                    # Criar cliente (EQUIPE não tem quarto)
                    customer = Customers(
                        nome=nome,
                        nickname=nickname,
                        tipo=CustomerTipo.EQUIPE,
                        quarto=None,  # Equipe não tem quarto
                        nome_pai=str(nome_pai).strip() if nome_pai else None,
                        nome_mae=str(nome_mae).strip() if nome_mae else None,
                        saldo=balance,
                        import_batch_id=batch.id,
                        created_by_id=user.id
                    )

                    db.add(customer)
                    db.flush()

                    imported_count += 1

                except Exception as e:
                    error_count += 1
                    errors_detail.append(f"Linha {row_num} (EQUIPE): Erro ao importar '{nome}': {str(e)}")
                    continue

    # Atualizar estatísticas do batch
    batch.imported_count = imported_count
    batch.skipped_count = skipped_count
    batch.error_count = error_count
    batch.status = "completed"
    db.commit()

    # 🆕 AUDITORIA: Registrar importação
    from app.services.audit import AuditService
    from app.models_audit import AuditAction

    audit = AuditService(db)
    audit.log_system_action(
        action=AuditAction.IMPORT,
        created_by_id=user.id,
        entity_type="customer_batch",
        entity_id=batch.id,
        new_values={
            "filename": filename,
            "imported_count": imported_count,
            "skipped_count": skipped_count,
            "error_count": error_count,
            "batch_id": batch.id
        },
        description=f"Importação de clientes: {imported_count} cliente(s) importado(s) de '{filename}' por {user.username}"
    )

    return {
        "success": True,
        "batch_id": batch.id,
        "imported": imported_count,
        "skipped": skipped_count,
        "errors": error_count,
        "errors_detail": errors_detail[:20]  # Limitar a 20 erros
    }


async def import_customers_from_upload(
    upload_file: UploadFile,
    db: Session,
    created_by_username: str = "admin"
) -> dict:
    """
    Importa clientes de um arquivo Excel enviado via upload
    """
    import tempfile

    # Salvar arquivo temporariamente
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        content = await upload_file.read()
        tmp_file.write(content)
        tmp_path = tmp_file.name

    try:
        result = import_customers_from_excel_file(tmp_path, db, created_by_username)
        return result
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def rollback_customer_batch(
    batch_id: int,
    db: Session,
    rolled_back_by_username: str
) -> dict:
    """
    Faz rollback de uma importação de clientes, deletando todos os clientes importados naquele batch
    """
    from app.models import Customers, CustomerImportBatch, BalanceTransaction, Sale
    from app.core.timezone import get_now

    batch = db.query(CustomerImportBatch).filter(CustomerImportBatch.id == batch_id).first()

    if not batch:
        return {"success": False, "error": "Batch de importação não encontrado"}

    if batch.status == "rolled_back":
        return {"success": False, "error": "Este batch já foi revertido anteriormente"}

    user = db.query(SystemUser).filter(SystemUser.username == rolled_back_by_username).first()

    if not user:
        return {"success": False, "error": "Usuário não encontrado"}

    customers = db.query(Customers).filter(Customers.import_batch_id == batch_id).all()

    if not customers:
        return {"success": False, "error": "Nenhum cliente encontrado para este batch"}

    # Verificar se algum cliente tem transações (vendas ou saldo)
    customers_with_transactions = []
    for customer in customers:
        sales_count = db.query(Sale).filter(Sale.customer_id == customer.id).count()
        balance_trans_count = db.query(BalanceTransaction).filter(BalanceTransaction.customer_id == customer.id).count()
        total_transactions = sales_count + balance_trans_count

        if total_transactions > 0:
            customers_with_transactions.append({
                "id": customer.id,
                "nome": customer.nome,
                "nickname": customer.nickname,
                "sales_count": sales_count,
                "balance_transactions_count": balance_trans_count,
                "total_transactions": total_transactions
            })

    if customers_with_transactions:
        return {
            "success": False,
            "error": "Não é possível reverter: alguns clientes já têm transações",
            "customers_with_transactions": customers_with_transactions[:5]
        }

    # Deletar clientes
    deleted_count = 0
    deleted_customers = []
    for customer in customers:
        deleted_customers.append({
            "id": customer.id,
            "nome": customer.nome,
            "nickname": customer.nickname,
            "tipo": customer.tipo.value,
            "saldo": customer.saldo
        })
        db.delete(customer)
        deleted_count += 1

    # Atualizar batch
    batch.status = "rolled_back"
    batch.rolled_back_at = get_now()
    batch.rolled_back_by_id = user.id

    db.commit()

    # 🆕 AUDITORIA: Registrar rollback
    from app.services.audit import AuditService
    from app.models_audit import AuditAction

    audit = AuditService(db)
    audit.log_system_action(
        action=AuditAction.ROLLBACK,
        created_by_id=user.id,
        entity_type="customer_batch",
        entity_id=batch_id,
        old_values={
            "status": "completed",
            "customers_count": deleted_count,
            "customers": deleted_customers[:10]
        },
        new_values={"status": "rolled_back"},
        description=f"Rollback de importação de clientes: {deleted_count} cliente(s) deletado(s) do batch #{batch_id} (arquivo: {batch.filename}) por {user.username}"
    )

    return {
        "success": True,
        "batch_id": batch_id,
        "deleted_count": deleted_count,
        "rolled_back_by": rolled_back_by_username,
        "rolled_back_at": batch.rolled_back_at.isoformat()
    }


def get_customer_import_batches_list(
    db: Session,
    skip: int = 0,
    limit: int = 50,
    include_rolled_back: bool = True
) -> list:
    """
    Lista todos os batches de importação de clientes
    """
    from app.models import Customers, CustomerImportBatch

    query = db.query(CustomerImportBatch)

    if not include_rolled_back:
        query = query.filter(CustomerImportBatch.status != "rolled_back")

    batches = query.order_by(CustomerImportBatch.created_at.desc()).offset(skip).limit(limit).all()

    result = []
    for batch in batches:
        customers_count = db.query(Customers).filter(Customers.import_batch_id == batch.id).count()

        result.append({
            "id": batch.id,
            "filename": batch.filename,
            "imported_count": batch.imported_count,
            "skipped_count": batch.skipped_count,
            "error_count": batch.error_count,
            "current_customers_count": customers_count,
            "status": batch.status,
            "created_at": batch.created_at.isoformat(),
            "created_by": batch.created_by.username if batch.created_by else None,
            "rolled_back_at": batch.rolled_back_at.isoformat() if batch.rolled_back_at else None,
            "rolled_back_by": batch.rolled_back_by.username if batch.rolled_back_by else None,
            "can_rollback": batch.status == "completed" and customers_count > 0
        })

    return result


