# endpoints/customers.py
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db
from app.core.dependencies import get_current_user, get_current_active_admin
from app.repositories import CustomerRepository
from app.models import (
    SystemUser, Customers, CustomerTipo, BalanceTransaction, Sale,
    CustomerImportBatch, CustomerImportBatchItem
)
from app import schemas
from app.services.audit import AuditService, get_changed_fields
from app.models_audit import AuditAction

router = APIRouter(prefix="/customers", tags=["customers"])


# ============================================
# CRUD de Customers
# ============================================

@router.post("", response_model=schemas.CustomerResponse)
def create_customer(
        customer: schemas.CustomerCreate,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)  # ← Apenas ADMIN
):
    customer_repo = CustomerRepository(db)

    # Verificar se nickname já existe
    if customer_repo.nickname_exists(customer.nickname):
        raise HTTPException(status_code=409, detail="Nickname já existe")

    # Criar novo cliente
    db_customer = Customers(
        nome=customer.nome,
        nickname=customer.nickname,
        quarto=customer.quarto,
        saldo=customer.saldo or 0.0,
        tipo=customer.tipo or CustomerTipo.ACAMPANTE,
        nome_pai=customer.nome_pai,
        nome_mae=customer.nome_mae,
        informacoes_contato=customer.informacoes_contato,
        created_by_id=current_user.id  # ← Registra quem criou
    )

    created_customer = customer_repo.create(db_customer)

    # 🆕 AUDITORIA: Registrar criação
    audit = AuditService(db)
    audit.log_customer_action(
        customer_id=created_customer.id,
        action=AuditAction.CREATE,
        created_by_id=current_user.id,
        new_values={
            "nome": created_customer.nome,
            "nickname": created_customer.nickname,
            "tipo": created_customer.tipo.value,
            "saldo": created_customer.saldo,
            "quarto": created_customer.quarto,
            "informacoes_contato": created_customer.informacoes_contato
        },
        description=f"Cliente criado por {current_user.username}"
    )

    return created_customer



@router.get("", response_model=List[schemas.CustomerResponse])
def read_customers(
        skip: int = 0,
        limit: int = 100,
        nome: Optional[str] = Query(None, description="Buscar por nome"),
        nickname: Optional[str] = Query(None, description="Buscar por ou nickname"),
        tipo: Optional[CustomerTipo] = Query(None, description="Filtrar por tipo"),
        customer_id: Optional[int] = Query(None, description="Buscar por ID"),
        is_active: Optional[bool] = Query(None, description="Filtrar por status ativo/inativo"),
        db: Session = Depends(get_db)
):
    """Lista todos os clientes com filtros opcionais"""
    customer_repo = CustomerRepository(db)

    if customer_id is not None:
        customer = customer_repo.get_by_id(customer_id)
        if customer is None:
            return []
        if is_active is not None and customer.is_active != is_active:
            return []
        return [customer]

    query = db.query(Customers)

    if is_active is not None:
        query = query.filter(Customers.is_active == is_active)
    if nome:
        query = query.filter(
            or_(Customers.nome.ilike(f"%{nome}%"), Customers.nickname.ilike(f"%{nome}%"))
        )
    if nickname:
        query = query.filter(Customers.nickname.ilike(f"%{nickname}%"))
    if tipo:
        query = query.filter(Customers.tipo == tipo)

    customers = query.offset(skip).limit(limit).all()
    return customers


@router.get("/{customer_id}", response_model=schemas.CustomerResponse)
def read_customer(
        customer_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Busca um cliente por ID"""
    customer_repo = CustomerRepository(db)
    customer = customer_repo.get_by_id(customer_id)

    if customer is None:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    return customer


@router.put("/{customer_id}", response_model=schemas.CustomerResponse)
def update_customer(
        customer_id: int,
        customer_update: schemas.CustomerUpdate,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)  # ← Apenas ADMIN
):
    """Atualiza dados de um cliente"""
    from datetime import datetime, timezone

    customer_repo = CustomerRepository(db)
    customer = customer_repo.get_by_id(customer_id)

    if customer is None:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    # Verificar se novo nickname já existe
    if customer_update.nickname and customer_update.nickname != customer.nickname:
        if customer_repo.nickname_exists(customer_update.nickname, exclude_id=customer_id):
            raise HTTPException(
                status_code=400,
                detail="Nickname já existe"
            )

    # 🆕 AUDITORIA: Detectar mudanças
    update_data = customer_update.model_dump(exclude_unset=True)
    old_values, new_values = get_changed_fields(customer, update_data)

    # Atualizar campos
    for field, value in update_data.items():
        setattr(customer, field, value)

    # Registrar quem e quando atualizou
    customer.updated_by_id = current_user.id
    customer.updated_at = datetime.now(timezone.utc)

    updated_customer = customer_repo.update(customer)

    # 🆕 AUDITORIA: Registrar mudanças (só se houver mudanças)
    if old_values:
        audit = AuditService(db)
        audit.log_customer_action(
            customer_id=customer_id,
            action=AuditAction.UPDATE,
            created_by_id=current_user.id,
            old_values=old_values,
            new_values=new_values,
            description=f"Cliente atualizado por {current_user.username}"
        )

    return updated_customer


@router.delete("/{customer_id}")
def delete_customer(
        customer_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)  # ← Apenas ADMIN
):
    """
    Deleta um cliente (soft delete - apenas desativa).
    NUNCA remove o cliente do banco de dados para manter integridade referencial.
    """
    from datetime import datetime, timezone

    customer_repo = CustomerRepository(db)
    customer = customer_repo.get_by_id(customer_id)

    if customer is None:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    # Verificar se já está desativado
    if not customer.is_active:
        raise HTTPException(status_code=400, detail="Cliente já está desativado")

    # Verificar histórico para mensagem informativa
    has_sales = db.query(Sale).filter(Sale.customer_id == customer_id).first()
    has_balance_transactions = db.query(BalanceTransaction).filter(
        BalanceTransaction.customer_id == customer_id
    ).first()

    # Determinar motivo da desativação para a mensagem
    reasons = []
    if has_sales:
        reasons.append("vendas")
    if has_balance_transactions:
        reasons.append("transações de saldo")

    reason_text = " e ".join(reasons) if reasons else "manter histórico"

    # SEMPRE soft delete (desativa)
    customer.is_active = False
    customer.updated_by_id = current_user.id
    customer.updated_at = datetime.now(timezone.utc)
    customer_repo.update(customer)

    # 🆕 AUDITORIA: Registrar desativação
    audit = AuditService(db)
    audit.log_customer_action(
        customer_id=customer_id,
        action=AuditAction.DEACTIVATE,
        created_by_id=current_user.id,
        old_values={"is_active": True},
        new_values={"is_active": False},
        description=f"Cliente desativado por {current_user.username} ({reason_text})"
    )

    return {
        "message": f"Cliente desativado com sucesso ({reason_text})",
        "is_active": False
    }


# ============================================
# Gerenciamento de Saldo
# ============================================

@router.post("/{customer_id}/balance", response_model=schemas.BalanceOperationResponse)
def manage_balance(
        customer_id: int,
        operation: schemas.BalanceOperation,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)  # ← Apenas ADMIN para gerenciar saldo manualmente
):
    """
    Adiciona ou remove saldo de um cliente. 
    - transaction_type: 'credit' (adiciona) ou 'debit' (remove)
    """
    customer_repo = CustomerRepository(db)
    customer = customer_repo.get_by_id(customer_id)

    if customer is None:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    if operation.amount <= 0:
        raise HTTPException(status_code=400, detail="Valor deve ser positivo")

    # Atualizar saldo
    if operation.transaction_type == "credit":
        customer.saldo += operation.amount
    elif operation.transaction_type == "debit":
        # Verificar se pode debitar
        if not customer.can_purchase(operation.amount):
            raise HTTPException(
                status_code=400,
                detail=f"Saldo insuficiente. Saldo atual: R$ {customer.saldo:.2f}"
            )
        customer.saldo -= operation.amount
    else:
        raise HTTPException(status_code=400, detail="Tipo de transação inválido")

    customer_repo.update(customer)

    # Criar registro de transação
    balance_transaction = BalanceTransaction(
        customer_id=customer_id,
        created_by_id=current_user.id,  # ← Rastreia quem fez
        amount=operation.amount,
        transaction_type=operation.transaction_type,
        description=operation.description or (
            "Recarga de saldo" if operation.transaction_type == "credit" else "Débito manual"
        )
    )
    db.add(balance_transaction)
    db.commit()
    db.refresh(balance_transaction)

    return {
        "message": f"Saldo {'adicionado' if operation.transaction_type == 'credit' else 'debitado'} com sucesso",
        "customer_id": customer_id,
        "customer_nome": customer.nome,
        "novo_saldo": customer.saldo,
        "valor_operacao": operation.amount,
        "transaction_id": balance_transaction.id
    }


@router.get("/{customer_id}/balance-history", response_model=schemas.BalanceHistoryResponse)
def get_balance_history(
        customer_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Retorna o histórico de transações de saldo do cliente"""
    customer_repo = CustomerRepository(db)
    customer = customer_repo.get_by_id(customer_id)

    if customer is None:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    transactions = db.query(BalanceTransaction) \
        .filter(BalanceTransaction.customer_id == customer_id) \
        .order_by(BalanceTransaction.created_at.desc()) \
        .all()

    return {
        "customer_id": customer_id,
        "customer_nome": customer.nome,
        "saldo_atual": customer.saldo,
        "tipo": customer.tipo.value,
        "historico": transactions
    }


# ============================================
# Relatórios e Estatísticas
# ============================================

@router.get("/{customer_id}/sales-summary", response_model=schemas.CustomerSalesSummary)
def get_customer_sales_summary(
        customer_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Retorna resumo de vendas do cliente"""
    customer_repo = CustomerRepository(db)
    customer = customer_repo.get_by_id(customer_id)

    if customer is None:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    # Resumo de vendas
    sales_summary = db.query(
        func.count(Sale.id).label('total_vendas'),
        func.sum(Sale.total_amount).label('total_gasto'),
        func.avg(Sale.total_amount).label('gasto_medio')
    ).filter(Sale.customer_id == customer_id).first()

    return {
        "customer_id": customer_id,
        "customer_nome": customer.nome,
        "tipo": customer.tipo.value,
        "saldo_atual": customer.saldo,
        "total_vendas": sales_summary.total_vendas or 0,
        "total_gasto": float(sales_summary.total_gasto or 0),
        "gasto_medio": float(sales_summary.gasto_medio or 0)
    }


@router.get("/{customer_id}/purchases", response_model=schemas.CustomerPurchaseHistory)
def get_customer_purchases(
        customer_id: int,
        limit: int = Query(50, description="Número máximo de registros"),
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Retorna o histórico de compras do cliente"""
    customer_repo = CustomerRepository(db)
    customer = customer_repo.get_by_id(customer_id)

    if not customer:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    sales = db.query(Sale) \
        .filter(Sale.customer_id == customer_id) \
        .order_by(Sale.created_at.desc()) \
        .limit(limit) \
        .all()

    purchases = []
    for sale in sales:
        items = []
        for item in sale.items:
            items.append({
                "produto_id": item.produto_id,
                "produto_nome": item.produto.nome if item.produto else "Produto removido",
                "quantity": item.quantity,
                "unit_price": item.unit_price,
                "total_price": item.total_price,
            })
        purchases.append({
            "id": sale.id,
            "total_amount": sale.total_amount,
            "created_at": sale.created_at,
            "is_cancelled": sale.is_cancelled,
            "items": items,
        })

    return {
        "customer_id": customer_id,
        "customer_nome": customer.nome,
        "total_compras": len(purchases),
        "purchases": purchases,
    }


# ============================================
# Ativação de Cliente
# ============================================

@router.patch("/{customer_id}/activate", response_model=schemas.CustomerResponse)
def activate_customer(
        customer_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Reativa um cliente desativado"""
    customer_repo = CustomerRepository(db)
    customer = customer_repo.get_by_id(customer_id)

    if customer is None:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    if customer.is_active:
        raise HTTPException(status_code=400, detail="Cliente já está ativo")

    customer.is_active = True
    customer.updated_by_id = current_user.id
    customer.updated_at = datetime.now(timezone.utc)
    updated = customer_repo.update(customer)

    audit = AuditService(db)
    audit.log_customer_action(
        customer_id=customer_id,
        action=AuditAction.UPDATE,
        created_by_id=current_user.id,
        old_values={"is_active": False},
        new_values={"is_active": True},
        description=f"Cliente reativado por {current_user.username}"
    )

    return updated


# ============================================
# Importação via Excel - Clientes
# ============================================

@router.get("/template/download")
def download_customer_template(
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Gera e retorna o template Excel para importação de clientes"""
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["nome", "nickname", "quarto", "nome_pai", "nome_mae", "informacoes_contato", "saldo"]

    for sheet_title, tipo_label in [("ACAMPANTES", "acampante"), ("EQUIPE", "equipe")]:
        ws = wb.create_sheet(title=sheet_title)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        if tipo_label == "acampante":
            ws.append(["João Silva", "joao", "101", "Carlos Silva", "Ana Silva", "", 50.0])
        else:
            ws.append(["Maria Santos", "maria", "", "", "", "Coordenadora", 0.0])

        for col_letter, width in zip("ABCDEFG", [30, 20, 10, 25, 25, 30, 10]):
            ws.column_dimensions[col_letter].width = width

    # Remover sheet padrão vazio
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template_importacao_clientes.xlsx"'}
    )


@router.post("/import")
def import_customers_from_excel(
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Importa clientes em massa a partir de um arquivo Excel com abas ACAMPANTES e EQUIPE"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx ou .xls")

    contents = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo Excel inválido ou corrompido")

    customer_repo = CustomerRepository(db)
    audit = AuditService(db)

    imported_count = 0
    skipped_count = 0
    error_count = 0
    imported_ids = []
    errors = []

    # Mapeia nome de aba para tipo de cliente
    sheet_tipo_map = {
        "ACAMPANTES": CustomerTipo.ACAMPANTE,
        "EQUIPE": CustomerTipo.EQUIPE,
    }

    # Se o arquivo tem abas nomeadas usa elas, senão usa a primeira como ACAMPANTES
    sheets_to_process = []
    for sheet_name, tipo in sheet_tipo_map.items():
        if sheet_name in wb.sheetnames:
            sheets_to_process.append((wb[sheet_name], tipo))

    if not sheets_to_process:
        # Arquivo sem abas padrão: trata a primeira aba como ACAMPANTES
        sheets_to_process = [(wb.active, CustomerTipo.ACAMPANTE)]

    for ws, tipo in sheets_to_process:
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for row_num, row in enumerate(rows, start=2):
            if not any(row):
                continue

            nome = str(row[0]).strip() if row[0] is not None else ""
            nickname = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

            if not nome or not nickname:
                skipped_count += 1
                continue

            if customer_repo.nickname_exists(nickname):
                skipped_count += 1
                continue

            try:
                quarto = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
                nome_pai = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
                nome_mae = str(row[4]).strip() if len(row) > 4 and row[4] is not None else None
                informacoes_contato = str(row[5]).strip() if len(row) > 5 and row[5] is not None else None
                saldo = float(row[6]) if len(row) > 6 and row[6] is not None else 0.0
            except (ValueError, TypeError) as e:
                errors.append(f"Linha {row_num}: dados inválidos para '{nome}' - {e}")
                error_count += 1
                continue

            db_customer = Customers(
                nome=nome,
                nickname=nickname,
                quarto=quarto or None,
                saldo=saldo,
                tipo=tipo,
                nome_pai=nome_pai or None,
                nome_mae=nome_mae or None,
                informacoes_contato=informacoes_contato or None,
                created_by_id=current_user.id
            )
            db.add(db_customer)
            db.flush()
            imported_ids.append(db_customer.id)

            audit.log_customer_action(
                customer_id=db_customer.id,
                action=AuditAction.CREATE,
                created_by_id=current_user.id,
                new_values={"nome": nome, "nickname": nickname, "tipo": tipo.value, "saldo": saldo},
                description=f"Cliente importado via Excel por {current_user.username}"
            )
            imported_count += 1

    # Criar batch de importação
    batch = CustomerImportBatch(
        filename=file.filename,
        imported_count=imported_count,
        skipped_count=skipped_count,
        error_count=error_count,
        status="completed",
        created_by_id=current_user.id
    )
    db.add(batch)
    db.flush()

    for customer_id in imported_ids:
        db.add(CustomerImportBatchItem(batch_id=batch.id, customer_id=customer_id))

    db.commit()

    return {
        "batch_id": batch.id,
        "filename": file.filename,
        "statistics": {
            "imported": imported_count,
            "skipped": skipped_count,
            "errors": error_count,
        },
        "errors": errors,
        "message": f"{imported_count} cliente(s) importado(s) com sucesso"
    }


@router.get("/import/batches")
def list_customer_import_batches(
        skip: int = 0,
        limit: int = 50,
        include_rolled_back: bool = True,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Lista todos os batches de importação de clientes"""
    query = db.query(CustomerImportBatch)

    if not include_rolled_back:
        query = query.filter(CustomerImportBatch.status == "completed")

    batches = query.order_by(CustomerImportBatch.created_at.desc()).offset(skip).limit(limit).all()

    result = []
    for batch in batches:
        customer_ids = [item.customer_id for item in batch.items]
        has_activity = False
        if customer_ids:
            has_sales = db.query(Sale).filter(Sale.customer_id.in_(customer_ids)).first()
            has_transactions = db.query(BalanceTransaction).filter(
                BalanceTransaction.customer_id.in_(customer_ids)
            ).first()
            has_activity = bool(has_sales or has_transactions)

        result.append({
            "id": batch.id,
            "filename": batch.filename,
            "imported_count": batch.imported_count,
            "skipped_count": batch.skipped_count,
            "error_count": batch.error_count,
            "status": batch.status,
            "created_at": batch.created_at,
            "created_by": batch.created_by.username,
            "rolled_back_at": batch.rolled_back_at,
            "rolled_back_by": batch.rolled_back_by.username if batch.rolled_back_by else None,
            "can_rollback": batch.status == "completed" and not has_activity,
        })

    return {"batches": result, "total": len(result)}


@router.get("/import/batches/{batch_id}")
def get_customer_import_batch_details(
        batch_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Retorna detalhes de um batch de importação de clientes"""
    batch = db.query(CustomerImportBatch).filter(CustomerImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch não encontrado")

    customers = []
    for item in batch.items:
        c = item.customer
        if c:
            customers.append({
                "id": c.id,
                "nome": c.nome,
                "nickname": c.nickname,
                "tipo": c.tipo.value,
                "quarto": c.quarto,
                "saldo": c.saldo,
                "is_active": c.is_active,
            })

    return {
        "id": batch.id,
        "filename": batch.filename,
        "imported_count": batch.imported_count,
        "skipped_count": batch.skipped_count,
        "error_count": batch.error_count,
        "status": batch.status,
        "created_at": batch.created_at,
        "created_by": batch.created_by.username,
        "rolled_back_at": batch.rolled_back_at,
        "rolled_back_by": batch.rolled_back_by.username if batch.rolled_back_by else None,
        "customers": customers,
    }


@router.delete("/import/batches/{batch_id}")
def rollback_customer_import_batch(
        batch_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Reverte um batch de importação deletando os clientes importados"""
    batch = db.query(CustomerImportBatch).filter(CustomerImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch não encontrado")

    if batch.status == "rolled_back":
        raise HTTPException(status_code=400, detail="Batch já foi revertido")

    customer_ids = [item.customer_id for item in batch.items]

    if customer_ids:
        has_sales = db.query(Sale).filter(Sale.customer_id.in_(customer_ids)).first()
        has_transactions = db.query(BalanceTransaction).filter(
            BalanceTransaction.customer_id.in_(customer_ids)
        ).first()
        if has_sales or has_transactions:
            raise HTTPException(
                status_code=400,
                detail="Não é possível reverter: alguns clientes já possuem vendas ou transações"
            )

    deleted_count = 0
    for customer_id in customer_ids:
        customer = db.query(Customers).filter(Customers.id == customer_id).first()
        if customer:
            db.delete(customer)
            deleted_count += 1

    db.query(CustomerImportBatchItem).filter(
        CustomerImportBatchItem.batch_id == batch_id
    ).delete()

    batch.status = "rolled_back"
    batch.rolled_back_at = datetime.now(timezone.utc)
    batch.rolled_back_by_id = current_user.id

    db.commit()

    return {
        "message": f"Rollback concluído. {deleted_count} cliente(s) removido(s).",
        "deleted_count": deleted_count,
        "batch_id": batch_id,
    }


@router.get("/stats/negative-balance", response_model=List[schemas.CustomerResponse])
def get_customers_with_negative_balance(
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Lista clientes com saldo negativo"""
    customer_repo = CustomerRepository(db)
    return customer_repo.get_with_negative_balance()


@router.get("/stats/by-type")
def get_customers_stats_by_type(
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Estatísticas de clientes por tipo"""
    stats = db.query(
        Customers.tipo,
        func.count(Customers.id).label('total'),
        func.sum(Customers.saldo).label('saldo_total'),
        func.avg(Customers.saldo).label('saldo_medio')
    ).group_by(Customers.tipo).all()

    return [
        {
            "tipo": stat.tipo.value,
            "total_clientes": stat.total,
            "saldo_total": float(stat.saldo_total or 0),
            "saldo_medio": float(stat.saldo_medio or 0)
        }
        for stat in stats
    ]


# ============================================
# Auditoria
# ============================================

@router.get("/{customer_id}/history")
def get_customer_history(
        customer_id: int,
        limit: int = Query(50, description="Número máximo de registros"),
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """
    Retorna o histórico completo de mudanças do cliente.
    Mostra todas as ações realizadas (criação, edições, ativação/desativação).
    """
    # Verificar se cliente existe
    customer_repo = CustomerRepository(db)
    customer = customer_repo.get_by_id(customer_id)
    if not customer:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    # Buscar histórico de auditoria
    audit = AuditService(db)
    history = audit.get_customer_history(customer_id, limit=limit)

    # Formatar resposta
    return {
        "customer_id": customer_id,
        "customer_nome": customer.nome,
        "total_actions": len(history),
        "history": [
            {
                "id": log.id,
                "action": log.action.value,
                "created_at": log.created_at,
                "created_by": log.created_by.username,
                "old_values": log.old_values,
                "new_values": log.new_values,
                "description": log.description
            }
            for log in history
        ]
    }
