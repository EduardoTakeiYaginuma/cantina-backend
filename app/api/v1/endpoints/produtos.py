# endpoints/produtos.py
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db
from app.core.dependencies import get_current_user, get_current_active_admin
from app.repositories import ProdutoRepository
from app.models import SystemUser, Produto, SaleItem, Restock, ProductImportBatch, ProductImportBatchItem
from app import schemas
from app.services.audit import AuditService, get_changed_fields
from app.models_audit import AuditAction

router = APIRouter(prefix="/produtos", tags=["produtos"])


# ============================================
# CRUD de Produtos
# ============================================

@router.post("", response_model=schemas.ProdutoResponse)
def create_produto(
        produto: schemas.ProdutoCreate,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)  # ← Apenas ADMIN
):
    """Cria um novo produto"""
    produto_repo = ProdutoRepository(db)

    # Verificar se nome já existe
    if produto_repo.nome_exists(produto.nome):
        raise HTTPException(
            status_code=400,
            detail="Produto com este nome já existe"
        )

    # Criar produto
    db_produto = Produto(
        nome=produto.nome,
        valor=produto.valor,
        estoque=produto.estoque or 0,
        estoque_minimo=produto.estoque_minimo or 10,
        created_by_id=current_user.id  # ← Registra quem criou
    )

    created_produto = produto_repo.create(db_produto)

    # 🆕 AUDITORIA: Registrar criação
    audit = AuditService(db)
    audit.log_product_action(
        produto_id=created_produto.id,
        action=AuditAction.CREATE,
        created_by_id=current_user.id,
        new_values={
            "nome": created_produto.nome,
            "valor": created_produto.valor,
            "estoque": created_produto.estoque,
            "estoque_minimo": created_produto.estoque_minimo
        },
        description=f"Produto criado por {current_user.username}"
    )

    return created_produto


@router.get("", response_model=List[schemas.ProdutoResponse])
def read_produtos(
        skip: int = 0,
        limit: int = 100,
        search: Optional[str] = Query(None, description="Buscar por nome"),
        low_stock: Optional[bool] = Query(None, description="Filtrar produtos com estoque baixo"),
        active_only: Optional[bool] = Query(True, description="Apenas produtos ativos"),
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Lista todos os produtos com filtros opcionais"""
    produto_repo = ProdutoRepository(db)

    if search:
        produtos = produto_repo.search(search)
    elif low_stock:
        produtos = produto_repo.get_low_stock()
    elif active_only:
        produtos = produto_repo.get_active_products()
    else:
        produtos = produto_repo.get_all(skip=skip, limit=limit)

    return produtos


@router.get("/{produto_id}", response_model=schemas.ProdutoResponse)
def read_produto(
        produto_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Busca um produto por ID"""
    produto_repo = ProdutoRepository(db)
    produto = produto_repo.get_by_id(produto_id)

    if produto is None:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    return produto


@router.put("/{produto_id}", response_model=schemas.ProdutoResponse)
def update_produto(
        produto_id: int,
        produto_update: schemas.ProdutoUpdate,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)  # ← Apenas ADMIN
):
    """Atualiza dados de um produto"""
    from datetime import datetime, timezone

    produto_repo = ProdutoRepository(db)
    produto = produto_repo.get_by_id(produto_id)

    if produto is None:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    # Verificar se novo nome já existe
    if produto_update.nome and produto_update.nome != produto.nome:
        if produto_repo.nome_exists(produto_update.nome, exclude_id=produto_id):
            raise HTTPException(
                status_code=400,
                detail="Produto com este nome já existe"
            )

    # 🆕 AUDITORIA: Detectar mudanças
    update_data = produto_update.model_dump(exclude_unset=True)
    old_values, new_values = get_changed_fields(produto, update_data)

    # Atualizar campos
    for field, value in update_data.items():
        setattr(produto, field, value)

    # Registrar quem e quando atualizou
    produto.updated_by_id = current_user.id
    produto.updated_at = datetime.now(timezone.utc)

    updated_produto = produto_repo.update(produto)

    # 🆕 AUDITORIA: Registrar mudanças
    if old_values:
        audit = AuditService(db)

        # Se mudou o preço, registrar ação ESPECIAL de mudança de preço
        if "valor" in old_values:
            audit.log_product_action(
                produto_id=produto_id,
                action=AuditAction.PRICE_CHANGE,
                created_by_id=current_user.id,
                old_values={"valor": old_values["valor"]},
                new_values={"valor": new_values["valor"]},
                description=f"Preço alterado de R${old_values['valor']:.2f} para R${new_values['valor']:.2f} por {current_user.username}"
            )

        # Se mudou outras coisas, registrar como UPDATE
        other_changes = {k: v for k, v in old_values.items() if k != "valor"}
        if other_changes:
            audit.log_product_action(
                produto_id=produto_id,
                action=AuditAction.UPDATE,
                created_by_id=current_user.id,
                old_values=other_changes,
                new_values={k: new_values[k] for k in other_changes.keys()},
                description=f"Produto atualizado por {current_user.username}"
            )

    return updated_produto


@router.delete("/{produto_id}")
def delete_produto(
        produto_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)  # ← Apenas ADMIN
):
    """
    Deleta um produto (soft delete - apenas desativa).
    NUNCA remove o produto do banco de dados para manter integridade referencial.
    """
    from datetime import datetime, timezone

    produto_repo = ProdutoRepository(db)
    produto = produto_repo.get_by_id(produto_id)

    if produto is None:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    # Verificar se já está desativado
    if not produto.is_active:
        raise HTTPException(status_code=400, detail="Produto já está desativado")

    # Verificar histórico para mensagem informativa
    has_sales = db.query(SaleItem).filter(SaleItem.produto_id == produto_id).first()
    has_restocks = db.query(Restock).filter(Restock.produto_id == produto_id).first()

    # Determinar motivo da desativação para a mensagem
    reasons = []
    if has_sales:
        reasons.append("vendas")
    if has_restocks:
        reasons.append("reabastecimentos")

    reason_text = " e ".join(reasons) if reasons else "manter histórico"

    # SEMPRE soft delete (desativa)
    produto.is_active = False
    produto.updated_by_id = current_user.id
    produto.updated_at = datetime.now(timezone.utc)
    produto_repo.update(produto)

    # 🆕 AUDITORIA: Registrar desativação
    audit = AuditService(db)
    audit.log_product_action(
        produto_id=produto_id,
        action=AuditAction.DEACTIVATE,
        created_by_id=current_user.id,
        old_values={"is_active": True},
        new_values={"is_active": False},
        description=f"Produto desativado por {current_user.username} ({reason_text})"
    )

    return {
        "message": f"Produto desativado com sucesso ({reason_text})",
        "is_active": False
    }


# ============================================
# Gerenciamento de Estoque
# ============================================

@router.post("/{produto_id}/restock", response_model=schemas.RestockResponse)
def restock_produto(
        produto_id: int,
        restock_data: schemas.RestockCreate,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)  # ← Apenas ADMIN para reabastecer
):
    """Reabastece o estoque de um produto"""
    produto_repo = ProdutoRepository(db)
    produto = produto_repo.get_by_id(produto_id)

    if produto is None:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    if restock_data.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantidade deve ser positiva")

    # Atualizar estoque
    old_stock = produto.estoque
    produto = produto_repo.add_stock(produto_id, restock_data.quantity)

    # Criar registro de reabastecimento
    restock = Restock(
        produto_id=produto_id,
        created_by_id=current_user.id,  # ← Rastreia quem fez
        quantity=restock_data.quantity
    )
    db.add(restock)
    db.commit()
    db.refresh(restock)

    return {
        "message": "Estoque reabastecido com sucesso",
        "produto_id": produto_id,
        "produto_nome": produto.nome,
        "estoque_anterior": old_stock,
        "quantidade_adicionada": restock_data.quantity,
        "estoque_atual": produto.estoque,
        "restock_id": restock.id,
        "realizado_por": current_user.username
    }


@router.get("/{produto_id}/restock-history", response_model=schemas.RestockHistoryResponse)
def get_restock_history(
        produto_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Retorna o histórico de reabastecimentos do produto"""
    produto_repo = ProdutoRepository(db)
    produto = produto_repo.get_by_id(produto_id)

    if produto is None:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    restocks = db.query(Restock) \
        .filter(Restock.produto_id == produto_id) \
        .order_by(Restock.created_at.desc()) \
        .all()

    return {
        "produto_id": produto_id,
        "produto_nome": produto.nome,
        "estoque_atual": produto.estoque,
        "estoque_minimo": produto.estoque_minimo,
        "historico_reabastecimento": restocks
    }


# ============================================
# Relatórios e Estatísticas
# ============================================

@router.get("/{produto_id}/sales-stats", response_model=schemas.ProdutoSalesStats)
def get_produto_sales_stats(
        produto_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Retorna estatísticas de vendas do produto"""
    produto_repo = ProdutoRepository(db)
    produto = produto_repo.get_by_id(produto_id)

    if produto is None:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    # Estatísticas de vendas
    sales_stats = db.query(
        func.count(SaleItem.id).label('total_vendas'),
        func.sum(SaleItem.quantity).label('quantidade_vendida'),
        func.sum(SaleItem.total_price).label('receita_total')
    ).filter(SaleItem.produto_id == produto_id).first()

    return {
        "produto_id": produto_id,
        "produto_nome": produto.nome,
        "produto_valor": produto.valor,
        "estoque_atual": produto.estoque,
        "estoque_minimo": produto.estoque_minimo,
        "is_active": produto.is_active,
        "total_vendas": sales_stats.total_vendas or 0,
        "quantidade_vendida": sales_stats.quantidade_vendida or 0,
        "receita_total": float(sales_stats.receita_total or 0)
    }


@router.get("/stats/low-stock", response_model=List[schemas.ProdutoResponse])
def get_low_stock_produtos(
        threshold: Optional[int] = Query(None, description="Limite customizado de estoque baixo"),
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Lista produtos com estoque baixo"""
    produto_repo = ProdutoRepository(db)

    if threshold:
        return produto_repo.get_low_stock(threshold)
    else:
        # Usa o estoque_minimo de cada produto
        return db.query(Produto).filter(
            Produto.estoque <= Produto.estoque_minimo
        ).all()


@router.get("/stats/summary")
def get_produtos_summary(
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """Retorna resumo geral de produtos"""
    total_produtos = db.query(func.count(Produto.id)).scalar()
    produtos_ativos = db.query(func.count(Produto.id)).filter(Produto.is_active == True).scalar()
    produtos_estoque_baixo = db.query(func.count(Produto.id)).filter(
        Produto.estoque <= Produto.estoque_minimo
    ).scalar()
    valor_total_estoque = db.query(func.sum(Produto.valor * Produto.estoque)).scalar()

    return {
        "total_produtos": total_produtos,
        "produtos_ativos": produtos_ativos,
        "produtos_inativos": total_produtos - produtos_ativos,
        "produtos_estoque_baixo": produtos_estoque_baixo,
        "valor_total_estoque": float(valor_total_estoque or 0)
    }


# ============================================
# Importação via Excel - Produtos
# ============================================

@router.get("/template/download")
def download_product_template(
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Gera e retorna o template Excel para importação de produtos"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"

    headers = ["nome", "valor", "estoque", "estoque_minimo"]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Exemplo de linha
    ws.append(["Coca-Cola", 5.00, 100, 10])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template_produtos.xlsx"'}
    )


@router.get("/restock/template/download")
def download_restock_template(
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Gera e retorna o template Excel para importação de reabastecimento em massa"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reabastecimento"

    headers = ["nome_produto", "quantidade"]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.append(["Coca-Cola", 50])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template_reabastecimento_produtos.xlsx"'}
    )


@router.post("/import")
def import_products_from_excel(
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Importa produtos em massa a partir de um arquivo Excel"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx ou .xls")

    contents = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo Excel inválido ou corrompido")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    produto_repo = ProdutoRepository(db)
    audit = AuditService(db)

    imported_count = 0
    skipped_count = 0
    error_count = 0
    imported_ids = []
    errors = []

    for row_num, row in enumerate(rows, start=2):
        if not any(row):
            continue

        nome = str(row[0]).strip() if row[0] is not None else ""
        if not nome:
            skipped_count += 1
            continue

        try:
            valor = float(row[1]) if row[1] is not None else None
            if valor is None or valor <= 0:
                errors.append(f"Linha {row_num}: valor inválido para '{nome}'")
                error_count += 1
                continue

            estoque = int(row[2]) if row[2] is not None else 0
            estoque_minimo = int(row[3]) if row[3] is not None else 10
        except (ValueError, TypeError):
            errors.append(f"Linha {row_num}: dados inválidos para '{nome}'")
            error_count += 1
            continue

        if produto_repo.nome_exists(nome):
            skipped_count += 1
            continue

        db_produto = Produto(
            nome=nome,
            valor=valor,
            estoque=max(0, estoque),
            estoque_minimo=max(0, estoque_minimo),
            created_by_id=current_user.id
        )
        created = produto_repo.create(db_produto)
        imported_ids.append(created.id)

        audit.log_product_action(
            produto_id=created.id,
            action=AuditAction.CREATE,
            created_by_id=current_user.id,
            new_values={"nome": created.nome, "valor": created.valor, "estoque": created.estoque},
            description=f"Produto importado via Excel por {current_user.username}"
        )
        imported_count += 1

    # Criar batch de importação
    batch = ProductImportBatch(
        filename=file.filename,
        imported_count=imported_count,
        skipped_count=skipped_count,
        error_count=error_count,
        status="completed",
        created_by_id=current_user.id
    )
    db.add(batch)
    db.flush()

    for produto_id in imported_ids:
        db.add(ProductImportBatchItem(batch_id=batch.id, produto_id=produto_id))

    db.commit()

    return {
        "batch_id": batch.id,
        "filename": file.filename,
        "statistics": {
            "imported": imported_count,
            "skipped": skipped_count,
            "errors": error_count,
        },
        "errors": errors,
        "message": f"{imported_count} produto(s) importado(s) com sucesso"
    }


@router.post("/restock/import")
def import_restock_from_excel(
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Importa reabastecimento em massa a partir de um arquivo Excel"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx ou .xls")

    contents = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo Excel inválido ou corrompido")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    produto_repo = ProdutoRepository(db)

    processed = 0
    skipped = 0
    errors = []

    for row_num, row in enumerate(rows, start=2):
        if not any(row):
            continue

        nome_produto = str(row[0]).strip() if row[0] is not None else ""
        if not nome_produto:
            skipped += 1
            continue

        try:
            quantidade = int(row[1]) if row[1] is not None else 0
            if quantidade <= 0:
                errors.append(f"Linha {row_num}: quantidade inválida para '{nome_produto}'")
                skipped += 1
                continue
        except (ValueError, TypeError):
            errors.append(f"Linha {row_num}: quantidade inválida para '{nome_produto}'")
            skipped += 1
            continue

        produto = db.query(Produto).filter(
            Produto.nome.ilike(nome_produto),
            Produto.is_active == True
        ).first()

        if not produto:
            errors.append(f"Linha {row_num}: produto '{nome_produto}' não encontrado")
            skipped += 1
            continue

        produto_repo.add_stock(produto.id, quantidade)

        restock = Restock(
            produto_id=produto.id,
            created_by_id=current_user.id,
            quantity=quantidade
        )
        db.add(restock)
        processed += 1

    db.commit()

    return {
        "statistics": {"processed": processed, "skipped": skipped},
        "errors": errors,
        "message": f"{processed} produto(s) reabastecido(s) com sucesso"
    }


@router.get("/import/batches")
def list_import_batches(
        skip: int = 0,
        limit: int = 50,
        include_rolled_back: bool = True,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Lista todos os batches de importação de produtos"""
    query = db.query(ProductImportBatch)

    if not include_rolled_back:
        query = query.filter(ProductImportBatch.status == "completed")

    batches = query.order_by(ProductImportBatch.created_at.desc()).offset(skip).limit(limit).all()

    result = []
    for batch in batches:
        # can_rollback: apenas se nenhum produto do batch foi vendido
        produto_ids = [item.produto_id for item in batch.items]
        has_sales = False
        if produto_ids:
            has_sales = db.query(SaleItem).filter(SaleItem.produto_id.in_(produto_ids)).first() is not None

        result.append({
            "id": batch.id,
            "filename": batch.filename,
            "imported_count": batch.imported_count,
            "skipped_count": batch.skipped_count,
            "error_count": batch.error_count,
            "status": batch.status,
            "created_at": batch.created_at,
            "created_by": batch.created_by.username,
            "rolled_back_at": batch.rolled_back_at,
            "rolled_back_by": batch.rolled_back_by.username if batch.rolled_back_by else None,
            "can_rollback": batch.status == "completed" and not has_sales,
        })

    return {"batches": result, "total": len(result)}


@router.get("/import/batches/{batch_id}")
def get_import_batch_details(
        batch_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Retorna detalhes de um batch de importação incluindo produtos"""
    batch = db.query(ProductImportBatch).filter(ProductImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch não encontrado")

    products = []
    for item in batch.items:
        p = item.produto
        if p:
            products.append({
                "id": p.id,
                "nome": p.nome,
                "valor": p.valor,
                "estoque": p.estoque,
                "is_active": p.is_active,
            })

    return {
        "id": batch.id,
        "filename": batch.filename,
        "imported_count": batch.imported_count,
        "skipped_count": batch.skipped_count,
        "error_count": batch.error_count,
        "status": batch.status,
        "created_at": batch.created_at,
        "created_by": batch.created_by.username,
        "rolled_back_at": batch.rolled_back_at,
        "rolled_back_by": batch.rolled_back_by.username if batch.rolled_back_by else None,
        "products": products,
    }


@router.delete("/import/batches/{batch_id}")
def rollback_import_batch(
        batch_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_active_admin)
):
    """Reverte um batch de importação deletando os produtos importados"""
    batch = db.query(ProductImportBatch).filter(ProductImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch não encontrado")

    if batch.status == "rolled_back":
        raise HTTPException(status_code=400, detail="Batch já foi revertido")

    produto_ids = [item.produto_id for item in batch.items]

    has_sales = False
    if produto_ids:
        has_sales = db.query(SaleItem).filter(SaleItem.produto_id.in_(produto_ids)).first() is not None

    if has_sales:
        raise HTTPException(
            status_code=400,
            detail="Não é possível reverter: alguns produtos já foram vendidos"
        )

    deleted_count = 0
    for produto_id in produto_ids:
        produto = db.query(Produto).filter(Produto.id == produto_id).first()
        if produto:
            db.delete(produto)
            deleted_count += 1

    db.query(ProductImportBatchItem).filter(ProductImportBatchItem.batch_id == batch_id).delete()

    batch.status = "rolled_back"
    batch.rolled_back_at = datetime.now(timezone.utc)
    batch.rolled_back_by_id = current_user.id

    db.commit()

    return {
        "message": f"Rollback concluído. {deleted_count} produto(s) removido(s).",
        "deleted_count": deleted_count,
        "batch_id": batch_id,
    }


# ============================================
# Auditoria
# ============================================

@router.get("/{produto_id}/history")
def get_product_history(
        produto_id: int,
        limit: int = Query(50, description="Número máximo de registros"),
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """
    Retorna o histórico completo de mudanças do produto.
    Mostra todas as ações realizadas (criação, edições, mudanças de preço, ativação/desativação).
    """
    # Verificar se produto existe
    produto_repo = ProdutoRepository(db)
    produto = produto_repo.get_by_id(produto_id)
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    # Buscar histórico de auditoria
    audit = AuditService(db)
    history = audit.get_product_history(produto_id, limit=limit)

    # Formatar resposta
    return {
        "produto_id": produto_id,
        "produto_nome": produto.nome,
        "total_actions": len(history),
        "history": [
            {
                "id": log.id,
                "action": log.action.value,
                "created_at": log.created_at,
                "created_by": log.created_by.username,
                "old_values": log.old_values,
                "new_values": log.new_values,
                "description": log.description
            }
            for log in history
        ]
    }


@router.get("/{produto_id}/price-history")
def get_product_price_history(
        produto_id: int,
        db: Session = Depends(get_db),
        current_user: SystemUser = Depends(get_current_user)
):
    """
    Retorna o histórico de mudanças de preço do produto.
    Útil para rastrear ajustes de preço ao longo do tempo.
    """
    # Verificar se produto existe
    produto_repo = ProdutoRepository(db)
    produto = produto_repo.get_by_id(produto_id)
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    # Buscar apenas mudanças de preço
    audit = AuditService(db)
    price_changes = audit.get_price_changes(produto_id)

    # Formatar resposta
    return {
        "produto_id": produto_id,
        "produto_nome": produto.nome,
        "valor_atual": produto.valor,
        "total_price_changes": len(price_changes),
        "price_history": [
            {
                "id": log.id,
                "created_at": log.created_at,
                "created_by": log.created_by.username,
                "old_price": log.old_values.get("valor") if log.old_values else None,
                "new_price": log.new_values.get("valor") if log.new_values else None,
                "description": log.description
            }
            for log in price_changes
        ]
    }


