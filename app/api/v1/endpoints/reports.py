# endpoints/reports.py
"""
Endpoints para exportação de relatórios de vendas em Excel e PDF.
"""
from datetime import datetime, date
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, and_
from sqlalchemy.orm import Session

from app.core.dependencies import get_current_active_admin
from app.models import SystemUser, Sale, SaleItem, Customers, Produto
from database import get_db

router = APIRouter(prefix="/reports", tags=["reports"])


def _build_report_data(db: Session, date_from: Optional[date], date_to: Optional[date], include_cancelled: bool):
    """Busca e agrega os dados de vendas para os relatórios."""
    query = db.query(Sale)

    if not include_cancelled:
        query = query.filter(Sale.is_cancelled == False)

    filters = []
    if date_from:
        filters.append(func.date(Sale.created_at) >= date_from)
    if date_to:
        filters.append(func.date(Sale.created_at) <= date_to)
    if filters:
        query = query.filter(and_(*filters))

    sales = query.order_by(Sale.created_at.asc()).all()

    # Totais globais
    total_revenue = sum(float(s.total_amount) for s in sales)
    total_sales_count = len(sales)
    avg_ticket = total_revenue / total_sales_count if total_sales_count else 0.0
    unique_customers = len({s.customer_id for s in sales})

    # Vendas por cliente
    customer_map: dict = {}
    for sale in sales:
        cid = sale.customer_id
        if cid not in customer_map:
            customer_map[cid] = {
                "nome": sale.customer.nome,
                "nickname": sale.customer.nickname,
                "tipo": sale.customer.tipo.value,
                "count": 0,
                "total": 0.0,
            }
        customer_map[cid]["count"] += 1
        customer_map[cid]["total"] += float(sale.total_amount)

    customers_sorted = sorted(customer_map.values(), key=lambda x: x["total"], reverse=True)

    # Vendas por produto
    product_map: dict = {}
    for sale in sales:
        for item in sale.items:
            pid = item.produto_id
            if pid not in product_map:
                product_map[pid] = {
                    "nome": item.produto.nome,
                    "quantity": 0,
                    "revenue": 0.0,
                    "unit_price": float(item.unit_price),
                }
            product_map[pid]["quantity"] += item.quantity
            product_map[pid]["revenue"] += float(item.total_price)

    products_sorted = sorted(product_map.values(), key=lambda x: x["revenue"], reverse=True)

    # Lista detalhada de vendas
    detail_rows = []
    for sale in reversed(sales):  # mais recente primeiro
        for item in sale.items:
            detail_rows.append({
                "sale_id": sale.id,
                "date": sale.created_at.strftime("%d/%m/%Y %H:%M"),
                "customer": sale.customer.nome,
                "nickname": sale.customer.nickname,
                "product": item.produto.nome,
                "quantity": item.quantity,
                "unit_price": float(item.unit_price),
                "total_price": float(item.total_price),
                "sale_total": float(sale.total_amount),
                "cancelled": "Sim" if sale.is_cancelled else "Não",
            })

    return {
        "generated_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "period": {
            "from": date_from.strftime("%d/%m/%Y") if date_from else "Início",
            "to": date_to.strftime("%d/%m/%Y") if date_to else "Hoje",
        },
        "summary": {
            "total_revenue": total_revenue,
            "total_sales": total_sales_count,
            "avg_ticket": avg_ticket,
            "unique_customers": unique_customers,
        },
        "by_customer": customers_sorted,
        "by_product": products_sorted,
        "detail": detail_rows,
    }


# ============================================
# Exportação Excel
# ============================================

@router.get("/excel")
def export_excel(
        include_cancelled: bool = Query(False, description="Incluir vendas canceladas"),
        db: Session = Depends(get_db),
        current_admin: SystemUser = Depends(get_current_active_admin),
):
    """Exporta relatório completo de vendas em Excel (.xlsx) com múltiplas abas."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="Biblioteca openpyxl não está disponível no servidor.")

    data = _build_report_data(db, None, None, include_cancelled)

    wb = Workbook()

    # Paleta de cores
    HEADER_COLOR = "1E3A5F"
    ACCENT_COLOR = "2563EB"
    ALT_ROW = "F0F4FF"
    WHITE = "FFFFFF"
    LIGHT_GRAY = "F8FAFC"

    header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    title_font = Font(name="Calibri", bold=True, size=14, color=HEADER_COLOR)
    subtitle_font = Font(name="Calibri", size=10, color="64748B")
    value_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", bold=True, size=11)

    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    accent_fill = PatternFill("solid", fgColor=ACCENT_COLOR)
    alt_fill = PatternFill("solid", fgColor=ALT_ROW)
    white_fill = PatternFill("solid", fgColor=WHITE)
    light_fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    def _set_col_widths(ws, widths: list[int]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_header_row(ws, row: int, cols: list[str]):
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=j, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        ws.row_dimensions[row].height = 20

    # ── Aba 1: Resumo ─────────────────────────────────────────────────────────

    ws1 = wb.active
    ws1.title = "Resumo"

    ws1.merge_cells("A1:D1")
    cell = ws1["A1"]
    cell.value = "Relatório de Vendas — Cantina"
    cell.font = title_font
    cell.alignment = center
    cell.fill = light_fill
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:D2")
    ws1["A2"] = f"Período: {data['period']['from']} até {data['period']['to']}   |   Gerado em: {data['generated_at']}"
    ws1["A2"].font = subtitle_font
    ws1["A2"].alignment = center
    ws1["A2"].fill = light_fill

    ws1.row_dimensions[3].height = 8  # espaço

    metrics = [
        ("Faturamento Total", f"R$ {data['summary']['total_revenue']:,.2f}"),
        ("Total de Vendas", str(data['summary']['total_sales'])),
        ("Ticket Médio", f"R$ {data['summary']['avg_ticket']:,.2f}"),
        ("Clientes Únicos", str(data['summary']['unique_customers'])),
    ]

    for i, (label, value) in enumerate(metrics):
        row = 4 + i
        ws1.row_dimensions[row].height = 22
        label_cell = ws1.cell(row=row, column=1, value=label)
        label_cell.font = bold_font
        label_cell.fill = light_fill
        label_cell.alignment = left
        label_cell.border = thin_border

        ws1.merge_cells(f"B{row}:D{row}")
        val_cell = ws1.cell(row=row, column=2, value=value)
        val_cell.font = Font(name="Calibri", bold=True, size=13, color=ACCENT_COLOR)
        val_cell.alignment = right
        val_cell.border = thin_border

    _set_col_widths(ws1, [30, 20, 20, 20])

    # ── Aba 2: Vendas por Cliente ──────────────────────────────────────────────

    ws2 = wb.create_sheet("Vendas por Cliente")
    headers2 = ["#", "Nome", "Nickname", "Tipo", "Nº Vendas", "Total (R$)", "% Faturamento"]
    _write_header_row(ws2, 1, headers2)

    total_rev = data["summary"]["total_revenue"] or 1
    for i, cust in enumerate(data["by_customer"], 1):
        row = i + 1
        pct = cust["total"] / total_rev * 100
        fill = alt_fill if i % 2 == 0 else white_fill
        row_data = [i, cust["nome"], cust["nickname"], cust["tipo"].capitalize(),
                    cust["count"], round(cust["total"], 2), round(pct, 2)]
        for j, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row, column=j, value=val)
            cell.font = value_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = right if j > 4 else left

    _set_col_widths(ws2, [5, 28, 18, 12, 12, 16, 16])

    # ── Aba 3: Vendas por Produto ──────────────────────────────────────────────

    ws3 = wb.create_sheet("Vendas por Produto")
    headers3 = ["#", "Produto", "Qtd. Vendida", "Preço Unit. (R$)", "Receita Total (R$)", "% Faturamento"]
    _write_header_row(ws3, 1, headers3)

    for i, prod in enumerate(data["by_product"], 1):
        row = i + 1
        pct = prod["revenue"] / total_rev * 100
        fill = alt_fill if i % 2 == 0 else white_fill
        row_data = [i, prod["nome"], prod["quantity"],
                    round(prod["unit_price"], 2), round(prod["revenue"], 2), round(pct, 2)]
        for j, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row, column=j, value=val)
            cell.font = value_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = right if j > 2 else left

    _set_col_widths(ws3, [5, 28, 14, 18, 20, 16])

    # ── Aba 4: Detalhamento de Vendas ─────────────────────────────────────────

    ws4 = wb.create_sheet("Detalhamento")
    headers4 = ["Venda ID", "Data/Hora", "Cliente", "Nickname", "Produto",
                "Qtd.", "Preço Unit. (R$)", "Total Item (R$)", "Total Venda (R$)", "Cancelado"]
    _write_header_row(ws4, 1, headers4)

    for i, row_d in enumerate(data["detail"], 1):
        row = i + 1
        fill = alt_fill if i % 2 == 0 else white_fill
        row_data = [
            row_d["sale_id"], row_d["date"], row_d["customer"], row_d["nickname"],
            row_d["product"], row_d["quantity"], round(row_d["unit_price"], 2),
            round(row_d["total_price"], 2), round(row_d["sale_total"], 2), row_d["cancelled"],
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws4.cell(row=row, column=j, value=val)
            cell.font = value_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = right if j in (6, 7, 8, 9) else left

    _set_col_widths(ws4, [10, 18, 25, 15, 25, 7, 16, 16, 16, 11])

    # Salvar e retornar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_vendas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ============================================
# Exportação PDF
# ============================================

@router.get("/pdf")
def export_pdf(
        include_cancelled: bool = Query(False, description="Incluir vendas canceladas"),
        db: Session = Depends(get_db),
        current_admin: SystemUser = Depends(get_current_active_admin),
):
    """Exporta relatório completo de vendas em PDF."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="Biblioteca reportlab não está disponível no servidor.")

    data = _build_report_data(db, None, None, include_cancelled)

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    DARK_BLUE = colors.HexColor("#1E3A5F")
    ACCENT_BLUE = colors.HexColor("#2563EB")
    LIGHT_BLUE = colors.HexColor("#EFF6FF")
    LIGHT_GRAY = colors.HexColor("#F8FAFC")
    ALT_ROW = colors.HexColor("#F0F4FF")
    MID_GRAY = colors.HexColor("#64748B")
    WHITE = colors.white

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 fontSize=18, textColor=DARK_BLUE,
                                 spaceAfter=4, fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               fontSize=9, textColor=MID_GRAY, spaceAfter=12)
    section_style = ParagraphStyle("section", parent=styles["Heading2"],
                                   fontSize=12, textColor=DARK_BLUE,
                                   spaceBefore=14, spaceAfter=6, fontName="Helvetica-Bold")

    def _table_style(has_alt=True):
        base = [
            ("BACKGROUND", (0, 0), (-1, 0), DARK_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUND", (0, 1), (-1, -1), [WHITE, ALT_ROW] if has_alt else [WHITE]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ROWHEIGHT", (0, 0), (-1, -1), 16),
        ]
        return TableStyle(base)

    story = []

    # Título
    story.append(Paragraph("Relatório de Vendas — Cantina", title_style))
    story.append(Paragraph(
        f"Período: {data['period']['from']} até {data['period']['to']}   ·   "
        f"Gerado em: {data['generated_at']}",
        sub_style
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=ACCENT_BLUE))
    story.append(Spacer(1, 10))

    # Resumo em cards (tabela de 4 colunas)
    summary_data = [
        ["Faturamento Total", "Total de Vendas", "Ticket Médio", "Clientes Únicos"],
        [
            f"R$ {data['summary']['total_revenue']:,.2f}",
            str(data['summary']['total_sales']),
            f"R$ {data['summary']['avg_ticket']:,.2f}",
            str(data['summary']['unique_customers']),
        ],
    ]
    summary_table = Table(summary_data, colWidths=[6.5 * cm] * 4)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), LIGHT_GRAY),
        ("BACKGROUND", (0, 1), (-1, 1), LIGHT_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), MID_GRAY),
        ("TEXTCOLOR", (0, 1), (-1, 1), ACCENT_BLUE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT", (0, 0), (-1, -1), 22),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 12))

    # Vendas por Cliente
    story.append(Paragraph("Vendas por Cliente", section_style))
    total_rev = data["summary"]["total_revenue"] or 1
    cust_headers = ["#", "Nome", "Nickname", "Tipo", "Nº Vendas", "Total (R$)", "% Faturamento"]
    cust_rows = [cust_headers]
    for i, c in enumerate(data["by_customer"], 1):
        pct = c["total"] / total_rev * 100
        cust_rows.append([
            str(i), c["nome"], c["nickname"], c["tipo"].capitalize(),
            str(c["count"]), f"R$ {c['total']:,.2f}", f"{pct:.1f}%",
        ])
    cust_col_w = [1.0 * cm, 5.5 * cm, 3.5 * cm, 2.5 * cm, 2.5 * cm, 3.5 * cm, 3.5 * cm]
    cust_table = Table(cust_rows, colWidths=cust_col_w, repeatRows=1)
    cust_table.setStyle(_table_style())
    story.append(cust_table)
    story.append(Spacer(1, 10))

    # Vendas por Produto
    story.append(Paragraph("Vendas por Produto", section_style))
    prod_headers = ["#", "Produto", "Qtd. Vendida", "Preço Unit. (R$)", "Receita Total (R$)", "% Faturamento"]
    prod_rows = [prod_headers]
    for i, p in enumerate(data["by_product"], 1):
        pct = p["revenue"] / total_rev * 100
        prod_rows.append([
            str(i), p["nome"], str(p["quantity"]),
            f"R$ {p['unit_price']:,.2f}", f"R$ {p['revenue']:,.2f}", f"{pct:.1f}%",
        ])
    prod_col_w = [1.0 * cm, 6.5 * cm, 3.0 * cm, 4.0 * cm, 4.5 * cm, 3.5 * cm]
    prod_table = Table(prod_rows, colWidths=prod_col_w, repeatRows=1)
    prod_table.setStyle(_table_style())
    story.append(prod_table)

    # Detalhamento de Vendas
    if data["detail"]:
        from reportlab.platypus import PageBreak
        story.append(PageBreak())
        story.append(Paragraph("Detalhamento de Vendas", section_style))
        story.append(Paragraph(
            "Uma linha por item vendido, agrupado por venda.",
            sub_style,
        ))

        det_headers = ["ID", "Data/Hora", "Cliente", "Produto", "Qtd.", "Unit. (R$)", "Item (R$)", "Venda (R$)"]
        det_rows = [det_headers]
        for row_d in data["detail"]:
            det_rows.append([
                str(row_d["sale_id"]),
                row_d["date"],
                row_d["customer"],
                row_d["product"],
                str(row_d["quantity"]),
                f"R$ {row_d['unit_price']:,.2f}",
                f"R$ {row_d['total_price']:,.2f}",
                f"R$ {row_d['sale_total']:,.2f}",
            ])

        det_col_w = [1.2 * cm, 3.8 * cm, 5.5 * cm, 5.5 * cm, 1.5 * cm, 3.5 * cm, 3.5 * cm, 3.5 * cm]
        det_table = Table(det_rows, colWidths=det_col_w, repeatRows=1)
        det_table.setStyle(_table_style())
        story.append(det_table)

    doc.build(story)
    output.seek(0)

    filename = f"relatorio_vendas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
