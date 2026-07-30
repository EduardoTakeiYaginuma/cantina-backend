"""
Script para gerar o template Excel de cadastro de produtos.

Uso:
    python templates/generate_template.py

Gera o arquivo templates/produtos_template.xlsx com as abas:
  - Bebidas, Doces, Salgados: tabelas estruturadas para preenchimento
  - Consolidado: consulta Power Query (M) que agrega as 3 tabelas
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(HERE, "produtos_template.xlsx")

HEADERS = ["Fardos", "ItensPorFardo", "QtdAvulsa", "Nome", "PrecoUnitario"]

# Column widths (in characters)
COL_WIDTHS = {
    "Fardos": 10,
    "ItensPorFardo": 14,
    "QtdAvulsa": 12,
    "Nome": 30,
    "PrecoUnitario": 16,
}

# Styling colours
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CURRENCY_FORMAT = 'R$ #,##0.00'

# Number of pre-formatted data rows (empty, ready to fill)
DATA_ROWS = 50

# Power Query M expression embedded in the Consolidado sheet as a named connection
# (also saved as a plain text file for reference)
POWER_QUERY_M = """\
let
    // ── Bebidas ──────────────────────────────────────────────────────────────
    Bebidas = Excel.CurrentWorkbook(){[Name="tbBebidas"]}[Content],
    BebidasTiped = Table.TransformColumnTypes(Bebidas, {
        {"Fardos",        type number},
        {"ItensPorFardo", type number},
        {"QtdAvulsa",     type number},
        {"Nome",          type text},
        {"PrecoUnitario", type number}
    }),
    BebidasFilled = Table.TransformColumns(BebidasTiped, {
        {"Fardos",        each if _ = null then 0 else _, type number},
        {"ItensPorFardo", each if _ = null then 0 else _, type number},
        {"QtdAvulsa",     each if _ = null then 0 else _, type number}
    }),
    BebidasResult = Table.SelectColumns(
        Table.AddColumn(BebidasFilled, "categoria", each "bebida"),
        {"categoria", "Nome", "Fardos", "ItensPorFardo", "QtdAvulsa", "PrecoUnitario"}
    ),

    // ── Doces ────────────────────────────────────────────────────────────────
    Doces = Excel.CurrentWorkbook(){[Name="tbDoces"]}[Content],
    DocesTiped = Table.TransformColumnTypes(Doces, {
        {"Fardos",        type number},
        {"ItensPorFardo", type number},
        {"QtdAvulsa",     type number},
        {"Nome",          type text},
        {"PrecoUnitario", type number}
    }),
    DocesFilled = Table.TransformColumns(DocesTiped, {
        {"Fardos",        each if _ = null then 0 else _, type number},
        {"ItensPorFardo", each if _ = null then 0 else _, type number},
        {"QtdAvulsa",     each if _ = null then 0 else _, type number}
    }),
    DocesResult = Table.SelectColumns(
        Table.AddColumn(DocesFilled, "categoria", each "doce"),
        {"categoria", "Nome", "Fardos", "ItensPorFardo", "QtdAvulsa", "PrecoUnitario"}
    ),

    // ── Salgados ─────────────────────────────────────────────────────────────
    Salgados = Excel.CurrentWorkbook(){[Name="tbSalgados"]}[Content],
    SalgadosTiped = Table.TransformColumnTypes(Salgados, {
        {"Fardos",        type number},
        {"ItensPorFardo", type number},
        {"QtdAvulsa",     type number},
        {"Nome",          type text},
        {"PrecoUnitario", type number}
    }),
    SalgadosFilled = Table.TransformColumns(SalgadosTiped, {
        {"Fardos",        each if _ = null then 0 else _, type number},
        {"ItensPorFardo", each if _ = null then 0 else _, type number},
        {"QtdAvulsa",     each if _ = null then 0 else _, type number}
    }),
    SalgadosResult = Table.SelectColumns(
        Table.AddColumn(SalgadosFilled, "categoria", each "salgado"),
        {"categoria", "Nome", "Fardos", "ItensPorFardo", "QtdAvulsa", "PrecoUnitario"}
    ),

    // ── União e cálculo de quantidade ────────────────────────────────────────
    All = Table.Combine({BebidasResult, DocesResult, SalgadosResult}),
    WithQty = Table.AddColumn(All, "quantidade",
        each [Fardos] * [ItensPorFardo] + [QtdAvulsa], type number),

    // ── Colunas finais para importação ───────────────────────────────────────
    Final = Table.RenameColumns(
        Table.SelectColumns(WithQty, {"categoria", "Nome", "quantidade", "PrecoUnitario"}),
        {{"Nome", "nome"}, {"PrecoUnitario", "preco_unitario"}}
    ),

    // ── Remove linhas sem nome (linhas vazias da tabela) ─────────────────────
    Filtered = Table.SelectRows(Final, each [nome] <> null and [nome] <> "")
in
    Filtered
"""


def add_input_sheet(wb: Workbook, sheet_name: str, table_name: str) -> None:
    """Create a formatted input sheet with a structured Excel table."""
    ws = wb.create_sheet(title=sheet_name)

    # ── Write headers ────────────────────────────────────────────────────────
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Pre-format data rows ─────────────────────────────────────────────────
    # Row banding is handled by the Excel Table style (showRowStripes=True).
    # Only apply the currency number format to the PrecoUnitario column here.
    for row in range(2, DATA_ROWS + 2):
        ws.cell(row=row, column=5).number_format = CURRENCY_FORMAT

    # ── Column widths ────────────────────────────────────────────────────────
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[header]

    ws.row_dimensions[1].height = 20

    # ── Freeze header row ────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Data validations ─────────────────────────────────────────────────────
    numeric_cols = [1, 2, 3, 5]  # Fardos, ItensPorFardo, QtdAvulsa, PrecoUnitario
    data_range = f"2:{DATA_ROWS + 1}"

    for col_idx in numeric_cols:
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error="Este campo aceita somente números não-negativos.",
        )
        dv.sqref = f"{col_letter}2:{col_letter}{DATA_ROWS + 1}"
        ws.add_data_validation(dv)

    # Nome required (non-empty text): use custom formula
    nome_col = get_column_letter(4)
    dv_nome = DataValidation(
        type="textLength",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Nome obrigatório",
        error="O campo Nome não pode ficar em branco.",
    )
    dv_nome.sqref = f"{nome_col}2:{nome_col}{DATA_ROWS + 1}"
    ws.add_data_validation(dv_nome)

    # ── Create Excel Table ───────────────────────────────────────────────────
    last_col = get_column_letter(len(HEADERS))
    table_ref = f"A1:{last_col}{DATA_ROWS + 1}"
    tbl = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tbl.tableStyleInfo = style
    ws.add_table(tbl)


def add_consolidado_sheet(wb: Workbook) -> None:
    """Create the Consolidado sheet with instructions and Power Query placeholder."""
    ws = wb.create_sheet(title="Consolidado")

    CONS_HEADERS = ["categoria", "nome", "quantidade", "preco_unitario"]

    # ── Section title ────────────────────────────────────────────────────────
    title_cell = ws.cell(row=1, column=1, value="CONSOLIDADO — Dados para Importação")
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 22

    # ── Instructions block ───────────────────────────────────────────────────
    instructions = [
        "Como atualizar este consolidado:",
        "  1. Preencha as abas Bebidas, Doces e Salgados com os produtos.",
        "  2. Clique em Dados > Atualizar Tudo (ou Ctrl+Alt+F5) para atualizar a consulta Power Query.",
        "  3. Os dados abaixo serão preenchidos automaticamente com categoria, nome, quantidade e preco_unitario.",
        "  4. Para exportar: selecione a tabela → Arquivo > Salvar Como → CSV UTF-8.",
        "",
        "Nota: a quantidade é calculada como (Fardos × ItensPorFardo) + QtdAvulsa.",
        "      Células vazias nas colunas numéricas são tratadas como 0.",
    ]

    instr_fill = PatternFill("solid", fgColor="EBF3FB")
    for i, line in enumerate(instructions, start=2):
        cell = ws.cell(row=i, column=1, value=line)
        cell.fill = instr_fill
        cell.alignment = Alignment(wrap_text=True)
        if i == 2:
            cell.font = Font(bold=True, italic=True)
        ws.merge_cells(f"A{i}:D{i}")

    ws.row_dimensions[2].height = 16
    for r in range(3, 2 + len(instructions)):
        ws.row_dimensions[r].height = 14

    # ── Data headers ─────────────────────────────────────────────────────────
    header_row = 2 + len(instructions)
    for col_idx, header in enumerate(CONS_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 20

    # ── Column widths ────────────────────────────────────────────────────────
    cons_col_widths = {"categoria": 14, "nome": 30, "quantidade": 14, "preco_unitario": 16}
    for col_idx, header in enumerate(CONS_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = cons_col_widths[header]

    # ── Power Query note ─────────────────────────────────────────────────────
    note_row = header_row + 1
    note = (
        "← Esta área será preenchida automaticamente pelo Power Query após 'Atualizar Tudo'. "
        "Consulte o arquivo power_query_consolidado.m para a expressão M completa."
    )
    note_cell = ws.cell(row=note_row, column=1, value=note)
    note_cell.font = Font(italic=True, color="595959")
    note_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{note_row}:D{note_row}")
    ws.row_dimensions[note_row].height = 28

    ws.freeze_panes = f"A{header_row + 1}"


def write_power_query_file() -> None:
    """Save the Power Query M expression as a plain-text .m file for reference."""
    path = os.path.join(HERE, "power_query_consolidado.m")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(POWER_QUERY_M)
    print(f"  → Power Query M: {path}")


def write_consolidado_csv() -> None:
    """Create the CSV header file used as import template."""
    path = os.path.join(HERE, "consolidado.csv")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("categoria,nome,quantidade,preco_unitario\n")
    print(f"  → CSV header:    {path}")


def generate() -> None:
    wb = Workbook()

    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # Input sheets
    for sheet_name, table_name in [
        ("Bebidas", "tbBebidas"),
        ("Doces", "tbDoces"),
        ("Salgados", "tbSalgados"),
    ]:
        print(f"  → Creating sheet '{sheet_name}' (table: {table_name})")
        add_input_sheet(wb, sheet_name, table_name)

    # Consolidado sheet
    print("  → Creating sheet 'Consolidado'")
    add_consolidado_sheet(wb)

    wb.save(OUTPUT_PATH)
    print(f"  → Workbook saved: {OUTPUT_PATH}")

    write_power_query_file()
    write_consolidado_csv()


if __name__ == "__main__":
    print("Gerando template Excel de cadastro de produtos…")
    generate()
    print("Concluído.")
